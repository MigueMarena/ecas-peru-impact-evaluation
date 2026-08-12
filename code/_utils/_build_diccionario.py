# -*- coding: utf-8 -*-
"""
Genera Diccionario_Variables_Creadas.xlsx con 3 hojas:
- Diccionario: 17 campos técnicos y 259 variables creadas en los scripts 06-15.
  Cabecera en azul institucional del BID; filas coloreadas por grupo temático
  (13 colores pastel consistentes con el catálogo para policy makers).
- Leyenda: valores controlados de los campos categóricos.
- Pendientes: variables cuyo rol analítico o tratamiento quedó por confirmar.

Convenciones:
- Definiciones limpias, sin detalles técnicos (esos van en 'notas') ni abreviaciones.
- Campo 'notas' reescrito para eliminar jerga Stata y expandir acrónimos (ECA,
  BPA, ENA, ITT, LATE, etc.) usando la misma lógica que el catálogo público.
- Ortografía con tildes (años, Campaña, etc.).
- Columna vars_input_derivadas sin anotaciones de script.

Autor: Carlos Marena (asistido)
Fecha: 2026-04-20
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from _notas_publicas import BID_BLUE, GROUP_COLORS, clean_notas

HEADERS = [
    "grupo", "subgrupo", "script", "variable_creada", "tipo",
    "vars_input_encuesta", "vars_input_derivadas", "nivel_agregacion",
    "universo_filtros", "unidad_medida", "definicion", "time_frame",
    "disponibilidad", "rol_analitico", "familia_outcome",
    "cuestionario_ref", "notas"
]

ROWS = []

def add(**kw):
    ROWS.append({h: kw.get(h, "") for h in HEADERS})

# ============================================================================
# GRUPO: Características de la Observación (script 06)
# ============================================================================
add(grupo="Características de la Observación", subgrupo="Efectos Fijos", script="06",
    variable_creada="cod_rgn_PE", tipo="categórica",
    vars_input_encuesta="", vars_input_derivadas="nomb_rgn, prod_ECA_eval",
    nivel_agregacion="Productor",
    universo_filtros="Todos",
    unidad_medida="código",
    definicion="Bloque de diseño formado por la interacción entre la región y el producto a evaluar. Identifica los estratos de aleatorización.",
    time_frame="En el momento de la encuesta",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Covariable",
    familia_outcome="",
    cuestionario_ref="",
    notas="Efecto fijo absorbido con a(cod_rgn_PE) en todas las regresiones ITT/LATE.")

add(grupo="Características de la Observación", subgrupo="Efectos Fijos", script="06",
    variable_creada="cod_cpb", tipo="categórica",
    vars_input_encuesta="",
    vars_input_derivadas="nomb_rgn, nomb_prvnc, nomb_dstrt, nomb_ccpp",
    nivel_agregacion="Centro Poblado",
    universo_filtros="Todos",
    unidad_medida="código",
    definicion="Centro poblado, unidad de aleatorización del programa, codificado como grupo único de región-provincia-distrito-ccpp.",
    time_frame="En el momento de la encuesta",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Filtro",
    familia_outcome="",
    cuestionario_ref="",
    notas="Usada para clusterizar errores estándar con cl(cod_cpb).")

add(grupo="Características de la Observación", subgrupo="Tiempo de la Encuesta", script="06",
    variable_creada="mes_enc", tipo="categórica",
    vars_input_encuesta="fch_enc",
    vars_input_derivadas="",
    nivel_agregacion="Productor",
    universo_filtros="Todos",
    unidad_medida="código (1-12)",
    definicion="Mes calendario en que se tomó la encuesta al productor.",
    time_frame="En el momento de la encuesta",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Covariable",
    familia_outcome="",
    cuestionario_ref="",
    notas="Incluido como i.mes_enc en las regresiones para absorber estacionalidad.")

add(grupo="Características de la Observación", subgrupo="Exposición a la ECA", script="06",
    variable_creada="dias_iniLB_iniECA", tipo="continua",
    vars_input_encuesta="fch_enc",
    vars_input_derivadas="fch_ini_1aECA",
    nivel_agregacion="Productor",
    universo_filtros="post==0 (línea base)",
    unidad_medida="días",
    definicion="Días transcurridos entre el inicio de la ECA y la fecha en que se aplicó la encuesta de línea base. Sirve como proxy de exposición inicial al programa.",
    time_frame="Campaña agrícola de línea base (2020-2021)",
    disponibilidad="Solo línea base",
    rol_analitico="Outcome",
    familia_outcome="Test Falsificación",
    cuestionario_ref="",
    notas="Es la variable endógena D en el test de falsificación del script 16, con asig_ccpp como instrumento Z. Imputada a 0 cuando fch_ini_1aECA es missing.")

# ============================================================================
# GRUPO: Sociodemográficas del Productor (script 07)
# ============================================================================
add(grupo="Sociodemográficas del Productor", subgrupo="Edad", script="07",
    variable_creada="edad", tipo="continua",
    vars_input_encuesta="preg003_1",
    vars_input_derivadas="",
    nivel_agregacion="Productor o Jefe de Hogar",
    universo_filtros="post==0",
    unidad_medida="años",
    definicion="Edad del productor o del jefe de hogar.",
    time_frame="En el momento de la línea base",
    disponibilidad="Solo línea base",
    rol_analitico="Covariable",
    familia_outcome="",
    cuestionario_ref="Módulo I",
    notas="Se imputa del roster de personas del hogar cuando no hay coincidencia directa con el productor (reclink2 con score >= 0.8249).")

add(grupo="Sociodemográficas del Productor", subgrupo="Edad", script="07",
    variable_creada="edadsq", tipo="continua",
    vars_input_encuesta="preg003_1",
    vars_input_derivadas="edad",
    nivel_agregacion="Productor o Jefe de Hogar",
    universo_filtros="post==0",
    unidad_medida="años al cuadrado",
    definicion="Edad del productor o del jefe de hogar elevada al cuadrado.",
    time_frame="En el momento de la línea base",
    disponibilidad="Solo línea base",
    rol_analitico="Covariable",
    familia_outcome="",
    cuestionario_ref="Módulo I",
    notas="Captura no-linealidad en la edad.")

add(grupo="Sociodemográficas del Productor", subgrupo="Sexo", script="07",
    variable_creada="sexo", tipo="binaria",
    vars_input_encuesta="preg002",
    vars_input_derivadas="",
    nivel_agregacion="Productor o Jefe de Hogar",
    universo_filtros="post==0",
    unidad_medida="0/1",
    definicion="Sexo del productor o del jefe de hogar: 1 si es hombre, 0 si es mujer.",
    time_frame="En el momento de la línea base",
    disponibilidad="Solo línea base",
    rol_analitico="Covariable",
    familia_outcome="",
    cuestionario_ref="Módulo I",
    notas="")

add(grupo="Sociodemográficas del Productor", subgrupo="Educación", script="07",
    variable_creada="educ", tipo="continua",
    vars_input_encuesta="preg005_1, preg005_2",
    vars_input_derivadas="",
    nivel_agregacion="Productor o Jefe de Hogar",
    universo_filtros="post==0",
    unidad_medida="años",
    definicion="Años de educación formal completados por el productor o el jefe de hogar.",
    time_frame="En el momento de la línea base",
    disponibilidad="Solo línea base",
    rol_analitico="Covariable",
    familia_outcome="",
    cuestionario_ref="Módulo I, preguntas 005",
    notas="Mapeo del nivel educativo (preg005_2) y grado (preg005_1) a años de escolaridad; incluye recodificaciones manuales para casos especiales identificados en preg803_1 (ocupación: docentes, profesores).")

add(grupo="Sociodemográficas del Productor", subgrupo="Educación", script="07",
    variable_creada="nivedmax", tipo="categórica",
    vars_input_encuesta="preg005_1, preg005_2",
    vars_input_derivadas="",
    nivel_agregacion="Productor o Jefe de Hogar",
    universo_filtros="post==0",
    unidad_medida="código (0-5)",
    definicion="Máximo nivel educativo completado por el productor o el jefe de hogar.",
    time_frame="En el momento de la línea base",
    disponibilidad="Solo línea base",
    rol_analitico="Covariable",
    familia_outcome="",
    cuestionario_ref="Módulo I, preguntas 005",
    notas="Categorías: 0=No estudió; 1=Inicial; 2=Primaria; 3=Secundaria; 4=Superior No Universitaria; 5=Superior Universitaria.")

add(grupo="Sociodemográficas del Productor", subgrupo="Lengua e Identidad", script="07",
    variable_creada="leng_mat", tipo="categórica",
    vars_input_encuesta="preg007",
    vars_input_derivadas="",
    nivel_agregacion="Productor o Jefe de Hogar",
    universo_filtros="post==0",
    unidad_medida="código (1-4)",
    definicion="Lengua materna del productor o del jefe de hogar.",
    time_frame="En el momento de la línea base",
    disponibilidad="Solo línea base",
    rol_analitico="Covariable",
    familia_outcome="",
    cuestionario_ref="Módulo I, pregunta 007",
    notas="Categorías: 1=Castellano; 2=Quechua; 3=Otras nativas/originarias (aimara, ashaninka, awajun, otra nativa); 4=Otras (extranjeras, no sabe/no contesta).")

add(grupo="Sociodemográficas del Productor", subgrupo="Lengua e Identidad", script="07",
    variable_creada="iden_etn", tipo="categórica",
    vars_input_encuesta="preg008",
    vars_input_derivadas="",
    nivel_agregacion="Productor o Jefe de Hogar",
    universo_filtros="post==0",
    unidad_medida="código (1-5)",
    definicion="Autoidentificación étnica del productor o del jefe de hogar.",
    time_frame="En el momento de la línea base",
    disponibilidad="Solo línea base",
    rol_analitico="Covariable",
    familia_outcome="",
    cuestionario_ref="Módulo I, pregunta 008",
    notas="Categorías: 1=Mestizo; 2=Quechua; 3=Otro indígena/originario; 4=Otro grupo étnico (afrodescendiente, blanco); 5=No sabe/no contesta.")

# ============================================================================
# GRUPO: Hogar - Activos y Condiciones de Vida (script 08)
# ============================================================================
add(grupo="Hogar - Activos y Condiciones de Vida", subgrupo="Activos Agrícolas", script="08",
    variable_creada="ilogsact", tipo="índice",
    vars_input_encuesta="preg201a_01 a preg201a_20",
    vars_input_derivadas="",
    nivel_agregacion="Hogar",
    universo_filtros="post==0",
    unidad_medida="índice logarítmico (ponderado 1-5)",
    definicion="Índice logarítmico de sofisticación de los activos agrícolas del hogar.",
    time_frame="En el momento de la línea base",
    disponibilidad="Solo línea base",
    rol_analitico="Covariable",
    familia_outcome="",
    cuestionario_ref="Módulo II, preguntas 201a",
    notas="Construido como suma ponderada de 5 grupos de activos ordenados por su nivel tecnológico (pesos 1 a 5), aplicando ln(1+n) dentro de cada grupo. Grupos: n1=herramientas manuales; n2=tracción animal/básicos; n3=arado mecánico; n4=motobombas/tractores; n5=avanzados (invernaderos, sensores).")

add(grupo="Hogar - Activos y Condiciones de Vida", subgrupo="Condiciones de Vida", script="08",
    variable_creada="icondvid", tipo="índice",
    vars_input_encuesta="preg901, preg902, preg903, preg904a, preg904b, preg905-preg913, preg914*",
    vars_input_derivadas="mat_par_ade, mat_pis_ade, mat_tec_ade, hab_score, amb_norm, acc_agua_pot_viv, elec_viv, coc_gas_hog, acc_intnet_hog, san_mej_viv, tot_act_bsc_norm, veh_mov",
    nivel_agregacion="Hogar",
    universo_filtros="post==0",
    unidad_medida="índice (0-12)",
    definicion="Índice sintético de condiciones de vida del hogar.",
    time_frame="En el momento de la línea base",
    disponibilidad="Solo línea base",
    rol_analitico="Covariable",
    familia_outcome="",
    cuestionario_ref="Módulos IX (Vivienda) y II (Activos)",
    notas="Suma de 12 componentes binarios o normalizados: materiales adecuados (paredes, pisos, techos), hacinamiento, ambientes, acceso a agua potable, electricidad, gas, internet, saneamiento mejorado, activos básicos y vehículos.")

# Ingresos del hogar
add(grupo="Hogar - Activos y Condiciones de Vida", subgrupo="Ingresos del Hogar", script="08",
    variable_creada="imen_otr_hog", tipo="continua",
    vars_input_encuesta="preg818a, preg818b, preg819b, preg819c, preg821a, preg821b, preg822a, preg822b, preg823a",
    vars_input_derivadas="",
    nivel_agregacion="Hogar",
    universo_filtros="post==0",
    unidad_medida="soles (S/.)",
    definicion="Ingreso mensual promedio del hogar por otros conceptos (transferencias, alquileres, jubilaciones y similares).",
    time_frame="Últimos 12 meses (anualizado a mensual)",
    disponibilidad="Solo línea base",
    rol_analitico="Por definir",
    familia_outcome="",
    cuestionario_ref="Módulo VIII",
    notas="Calculado como el total anual dividido entre 12.")

add(grupo="Hogar - Activos y Condiciones de Vida", subgrupo="Ingresos del Hogar", script="08",
    variable_creada="imen_dep_hog", tipo="continua",
    vars_input_encuesta="preg802_*, preg804_*, preg805_*, preg806_*, preg808a_*, preg808b_*",
    vars_input_derivadas="",
    nivel_agregacion="Hogar",
    universo_filtros="post==0",
    unidad_medida="soles (S/.)",
    definicion="Ingreso mensual anualizado del hogar proveniente del trabajo dependiente (agrícola y no agrícola) de sus miembros.",
    time_frame="Últimos 12 meses (anualizado a mensual)",
    disponibilidad="Solo línea base",
    rol_analitico="Por definir",
    familia_outcome="",
    cuestionario_ref="Módulo VIII",
    notas="Convierte los ingresos reportados según la frecuencia declarada (diaria, semanal, quincenal, mensual, trimestral, semestral, anual) y los ajusta por los meses efectivamente trabajados. Suma sobre hasta 3 trabajos reportados por cada uno de los miembros del hogar.")

add(grupo="Hogar - Activos y Condiciones de Vida", subgrupo="Ingresos del Hogar", script="08",
    variable_creada="imen_indep_hog", tipo="continua",
    vars_input_encuesta="preg811_*, preg813_*, preg816_*",
    vars_input_derivadas="",
    nivel_agregacion="Hogar",
    universo_filtros="post==0",
    unidad_medida="soles (S/.)",
    definicion="Ingreso mensual anualizado del hogar proveniente del trabajo independiente de sus miembros.",
    time_frame="Últimos 12 meses (anualizado a mensual)",
    disponibilidad="Solo línea base",
    rol_analitico="Por definir",
    familia_outcome="",
    cuestionario_ref="Módulo VIII",
    notas="Corrección metodológica: cuando el ingreso declarado supera 10000 soles se asume frecuencia anual y se reexpresa en mensual. Suma sobre hasta 2 trabajos independientes por miembro.")

add(grupo="Hogar - Activos y Condiciones de Vida", subgrupo="Ingresos del Hogar", script="08",
    variable_creada="imen_hog", tipo="continua",
    vars_input_encuesta="",
    vars_input_derivadas="imen_dep_hog, imen_indep_hog, imen_otr_hog",
    nivel_agregacion="Hogar",
    universo_filtros="post==0",
    unidad_medida="soles (S/.)",
    definicion="Ingreso monetario mensual anualizado total del hogar: suma de ingresos dependientes, independientes y otros.",
    time_frame="Últimos 12 meses (anualizado a mensual)",
    disponibilidad="Solo línea base",
    rol_analitico="Por definir",
    familia_outcome="",
    cuestionario_ref="",
    notas="")

add(grupo="Hogar - Activos y Condiciones de Vida", subgrupo="Ingresos del Hogar", script="08",
    variable_creada="imen_per_cap", tipo="continua",
    vars_input_encuesta="",
    vars_input_derivadas="imen_hog, tot_miem",
    nivel_agregacion="Hogar",
    universo_filtros="post==0",
    unidad_medida="soles (S/.)",
    definicion="Ingreso monetario mensual anualizado per cápita del hogar.",
    time_frame="Últimos 12 meses (anualizado a mensual)",
    disponibilidad="Solo línea base",
    rol_analitico="Por definir",
    familia_outcome="",
    cuestionario_ref="",
    notas="Cociente entre imen_hog y el total de miembros del hogar.")

add(grupo="Hogar - Activos y Condiciones de Vida", subgrupo="Ingresos del Hogar", script="08",
    variable_creada="log_imen_per_cap", tipo="continua",
    vars_input_encuesta="",
    vars_input_derivadas="imen_per_cap",
    nivel_agregacion="Hogar",
    universo_filtros="post==0",
    unidad_medida="log soles",
    definicion="Logaritmo natural del ingreso monetario mensual anualizado per cápita del hogar.",
    time_frame="Últimos 12 meses (anualizado a mensual)",
    disponibilidad="Solo línea base",
    rol_analitico="Por definir",
    familia_outcome="",
    cuestionario_ref="",
    notas="")

# ============================================================================
# GRUPO: Hogar - Composición Demográfica (script 08)
# ============================================================================
add(grupo="Hogar - Composición Demográfica", subgrupo="Grupos Etarios", script="08",
    variable_creada="tot_miem_1564", tipo="conteo",
    vars_input_encuesta="preg003_1, preg002",
    vars_input_derivadas="tot_miem_H_1564, tot_miem_M_1564",
    nivel_agregacion="Hogar",
    universo_filtros="post==0",
    unidad_medida="personas",
    definicion="Total de miembros del hogar en edad de no dependencia (entre 15 y 64 años).",
    time_frame="En el momento de la línea base",
    disponibilidad="Solo línea base",
    rol_analitico="Covariable",
    familia_outcome="",
    cuestionario_ref="Módulo I (Personas)",
    notas="")

add(grupo="Hogar - Composición Demográfica", subgrupo="Grupos Etarios", script="08",
    variable_creada="tot_miem_depen", tipo="conteo",
    vars_input_encuesta="preg003_1",
    vars_input_derivadas="",
    nivel_agregacion="Hogar",
    universo_filtros="post==0",
    unidad_medida="personas",
    definicion="Total de miembros del hogar en edad de dependencia (menores de 15 años o con 65 años o más).",
    time_frame="En el momento de la línea base",
    disponibilidad="Solo línea base",
    rol_analitico="Covariable",
    familia_outcome="",
    cuestionario_ref="Módulo I (Personas)",
    notas="")

# ============================================================================
# GRUPO: Hogar - Servicios de Extensión Agraria (script 08)
# ============================================================================
_SEA_LBLS = [
    ("sea1",  "Análisis de suelos"),
    ("sea2",  "Técnicas de labranza de suelos"),
    ("sea3",  "Rotación de cultivos"),
    ("sea4",  "Técnicas de manejo de semillas"),
    ("sea5",  "Operación y mantenimiento de sistemas de riego"),
    ("sea6",  "Sistemas de riego tecnificado"),
    ("sea7",  "Prácticas adecuadas de riego"),
    ("sea8",  "Uso de abonos y fertilizantes"),
    ("sea9",  "Uso de plaguicidas"),
    ("sea10", "Uso de control biológico"),
    ("sea11", "Manejo integrado de plagas"),
    ("sea12", "Estándares de calidad de agua para riego"),
    ("sea13", "Buenas Prácticas Agrícolas"),
    ("sea14", "Producción orgánica"),
    ("sea15", "Instalación y manejo de pastos"),
    ("sea16", "Alimentación de animales de crianza"),
    ("sea17", "Mejoramiento genético de animales"),
    ("sea18", "Vacunas y medicamentos veterinarios"),
    ("sea19", "Prácticas de bioseguridad"),
    ("sea20", "Buenas prácticas pecuarias"),
    ("sea21", "Manipulación e higiene de alimentos"),
    ("sea22", "Almacenamiento de alimentos"),
    ("sea23", "Contaminación de alimentos"),
]
for i, (vname, tema) in enumerate(_SEA_LBLS, start=1):
    add(grupo="Hogar - Servicios de Extensión Agraria", subgrupo="Capacitaciones recibidas", script="08",
        variable_creada=vname, tipo="binaria",
        vars_input_encuesta=f"preg602{i}",
        vars_input_derivadas="",
        nivel_agregacion="Hogar",
        universo_filtros="post==0",
        unidad_medida="0/1",
        definicion=f"Algún miembro del hogar recibió capacitación en {tema} durante los últimos tres años.",
        time_frame="Últimos tres años",
        disponibilidad="Solo línea base",
        rol_analitico="Por definir",
        familia_outcome="",
        cuestionario_ref="Módulo VI, servicios de extensión",
        notas="")

add(grupo="Hogar - Servicios de Extensión Agraria", subgrupo="Dummies Agregadas", script="08",
    variable_creada="cap_suelo", tipo="binaria",
    vars_input_encuesta="preg6021-preg6024",
    vars_input_derivadas="sea1, sea2, sea3, sea4",
    nivel_agregacion="Hogar",
    universo_filtros="post==0",
    unidad_medida="0/1",
    definicion="Algún miembro del hogar recibió capacitación en prácticas para minimizar la degradación del suelo.",
    time_frame="Últimos tres años",
    disponibilidad="Solo línea base",
    rol_analitico="Por definir",
    familia_outcome="",
    cuestionario_ref="Módulo VI",
    notas="OR de sea1 a sea4.")

add(grupo="Hogar - Servicios de Extensión Agraria", subgrupo="Dummies Agregadas", script="08",
    variable_creada="cap_riego", tipo="binaria",
    vars_input_encuesta="preg6025-preg6027, preg60212",
    vars_input_derivadas="sea5, sea6, sea7, sea12",
    nivel_agregacion="Hogar",
    universo_filtros="post==0",
    unidad_medida="0/1",
    definicion="Algún miembro del hogar recibió capacitación en prácticas para mejorar el riego.",
    time_frame="Últimos tres años",
    disponibilidad="Solo línea base",
    rol_analitico="Por definir",
    familia_outcome="",
    cuestionario_ref="Módulo VI",
    notas="OR de sea5, sea6, sea7 y sea12.")

add(grupo="Hogar - Servicios de Extensión Agraria", subgrupo="Dummies Agregadas", script="08",
    variable_creada="cap_insum", tipo="binaria",
    vars_input_encuesta="preg6028-preg60211",
    vars_input_derivadas="sea8, sea9, sea10, sea11",
    nivel_agregacion="Hogar",
    universo_filtros="post==0",
    unidad_medida="0/1",
    definicion="Algún miembro del hogar recibió capacitación en prácticas para mejorar la aplicación de insumos.",
    time_frame="Últimos tres años",
    disponibilidad="Solo línea base",
    rol_analitico="Por definir",
    familia_outcome="",
    cuestionario_ref="Módulo VI",
    notas="OR de sea8 a sea11.")

# ============================================================================
# GRUPO: Predio y Parcelas (script 10)
# ============================================================================
add(grupo="Predio y Parcelas", subgrupo="Extensión", script="10",
    variable_creada="tot_has_prod", tipo="continua",
    vars_input_encuesta="preg101c, preg101c1",
    vars_input_derivadas="tot_has_pp",
    nivel_agregacion="Productor",
    universo_filtros="Todos",
    unidad_medida="hectáreas",
    definicion="Total de hectáreas gestionadas por el productor, sumando la superficie de todos sus predios.",
    time_frame="En el momento de la encuesta",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Covariable",
    familia_outcome="",
    cuestionario_ref="Módulo III (Extensión del predio)",
    notas="Agregación vía bys Codprod22 post: egen total(tot_has_pp).")

add(grupo="Predio y Parcelas", subgrupo="Tenencia", script="10",
    variable_creada="años_tenen_prod", tipo="continua",
    vars_input_encuesta="preg105",
    vars_input_derivadas="años_tenen_pp",
    nivel_agregacion="Productor",
    universo_filtros="Todos",
    unidad_medida="años",
    definicion="Años de tenencia del predio más antiguo del productor. Sirve como proxy de su experiencia.",
    time_frame="En el momento de la encuesta",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Covariable",
    familia_outcome="",
    cuestionario_ref="Módulo III (Tenencia)",
    notas="Construida como el máximo sobre los predios del productor.")

add(grupo="Predio y Parcelas", subgrupo="Riego", script="10",
    variable_creada="riego_tec_prod", tipo="binaria",
    vars_input_encuesta="preg101a, preg103",
    vars_input_derivadas="riego_tec_pp",
    nivel_agregacion="Productor",
    universo_filtros="Parcela principal (preg101a==1)",
    unidad_medida="0/1",
    definicion="La parcela principal del productor cuenta con riego tecnificado.",
    time_frame="En el momento de la encuesta",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Covariable;Filtro",
    familia_outcome="",
    cuestionario_ref="Módulo III (Riego)",
    notas="Usada también como filtro para análisis de BPAs condicionadas a riego en los scripts 19 y 25A. Construida como el máximo sobre los predios del productor filtrando por parcela principal.")

# Gastos a nivel predio (_pp)
_GTOT_PP = [
    ("gtot_alq_pp",        "alquiler del predio o parcela (dinero + bienes + % producción valorizado)", "preg107a, preg107b, preg107c2 + vtot_cose_pp (de 09)"),
    ("gtot_alq_a_pp",      "alquiler del predio o parcela (dinero + bienes)",                          "preg107a, preg107c2"),
    ("gtot_alq_b_pp",      "alquiler del predio o parcela valorizado del pago como % de producción",   "preg107b × vtot_cose_pp (de 09)"),
    ("gtot_ases_tecn",     "asesoría técnica (abono, fertilización, riego, manejo de cultivos, combate de plagas)", "preg115b, preg115d"),
    ("gtot_plantones_pp",  "compra de plantones",                                   "preg115f5_*"),
    ("gtot_mo_act_agr",    "mano de obra en actividades agrícolas (jornales y tareas)", "preg116a1a, preg116a1b, preg116a2a, preg116a2b"),
    ("gtot_mo_ctrl_pla",   "mano de obra en prevención y control de plagas",        "preg116b1a, preg116b1b, preg116b2a, preg116b2b"),
    ("gtot_maq_pp",        "alquiler de maquinaria",                                "preg116c1, preg116c2"),
    ("gtot_fanim_pp",      "alquiler de fuerza animal",                             "preg116d1, preg116d2"),
    ("gtot_agua_riego_pp", "agua para riego",                                       "preg116ea5c, preg116eb5c, preg116ec5c"),
    ("gtot_mo_riego_pp",   "mano de obra en labores de riego",                      "preg116ea6, preg116eb6, preg116ec6"),
    ("gtot_met_mip_pp",    "métodos de manejo integrado de plagas",                 "preg117a, preg117b"),
    ("gtot_tpte_prod_pp",  "transporte de insumos y producción",                    "preg117c"),
    ("gtot_plagui_q_pp",   "plaguicidas (herbicidas, insecticidas, fungicidas)",    "preg117e"),
    ("gtot_abo_fert_q_pp", "abonos y fertilizantes químicos",                       "preg117g"),
    ("gtot_abo_org_pp",    "abono orgánico o natural",                              "preg117f"),
    ("gtot_otr_pp",        "otros conceptos de producción agrícola",                "preg117d"),
]
for vname, desc, pregs in _GTOT_PP:
    add(grupo="Predio y Parcelas", subgrupo="Costos - Predio", script="10",
        variable_creada=vname, tipo="continua",
        vars_input_encuesta=pregs,
        vars_input_derivadas="",
        nivel_agregacion="Parcela",
        universo_filtros="Todos",
        unidad_medida="soles (S/.)",
        definicion=f"Gasto total en {desc} del productor a nivel de parcela durante la campaña agrícola.",
        time_frame="Campaña agrícola",
        disponibilidad="Línea base y línea seguimiento",
        rol_analitico="Outcome",
        familia_outcome="Económicos",
        cuestionario_ref="Módulos de costos y tenencia",
        notas="")

add(grupo="Predio y Parcelas", subgrupo="Costos Operativos - Predio", script="10",
    variable_creada="gtot_oper_1_pp", tipo="continua",
    vars_input_encuesta="",
    vars_input_derivadas="gtot_alq_pp (= gtot_alq_a_pp + gtot_alq_b_pp), gtot_ases_tecn, gtot_plantones_pp, gtot_mo_act_agr, gtot_mo_ctrl_pla, gtot_maq_pp, gtot_fanim_pp, gtot_agua_riego_pp, gtot_mo_riego_pp, gtot_met_mip_pp, gtot_tpte_prod_pp, gtot_plagui_q_pp, gtot_abo_fert_q_pp, gtot_abo_org_pp, gtot_otr_pp",
    nivel_agregacion="Parcela",
    universo_filtros="Todos",
    unidad_medida="soles (S/.)",
    definicion="Gasto total en costos operativos a nivel parcela, incluyendo insumos químicos y orgánicos.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="",
    notas="Suma de los 15 componentes de costo a nivel parcela.")

add(grupo="Predio y Parcelas", subgrupo="Costos Operativos - Predio", script="10",
    variable_creada="gtot_oper_2_pp", tipo="continua",
    vars_input_encuesta="",
    vars_input_derivadas="gtot_alq_pp (= gtot_alq_a_pp + gtot_alq_b_pp), gtot_ases_tecn, gtot_plantones_pp, gtot_mo_act_agr, gtot_mo_ctrl_pla, gtot_maq_pp, gtot_fanim_pp, gtot_agua_riego_pp, gtot_mo_riego_pp, gtot_met_mip_pp, gtot_tpte_prod_pp, gtot_otr_pp",
    nivel_agregacion="Parcela",
    universo_filtros="Todos",
    unidad_medida="soles (S/.)",
    definicion="Gasto total en costos operativos a nivel parcela, excluyendo insumos químicos y orgánicos.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="",
    notas="Suma de 12 componentes, excluyendo gtot_plagui_q_pp, gtot_abo_fert_q_pp y gtot_abo_org_pp.")

# Agregaciones al productor
_GTOT_PROD = [
    ("gtot_alq_prod",        "gtot_alq_pp",        "alquiler (dinero + bienes + % producción valorizado)"),
    ("gtot_alq_a_prod",      "gtot_alq_a_pp",      "alquiler (componente: dinero + bienes)"),
    ("gtot_alq_b_prod",      "gtot_alq_b_pp",      "alquiler (componente: % de producción valorizado)"),
    ("gtot_ases_tecn_prod",  "gtot_ases_tecn",     "asesoría técnica"),
    ("gtot_plantones_prod",  "gtot_plantones_pp",  "compra de plantones"),
    ("gtot_mo_act_agr_prod", "gtot_mo_act_agr",    "mano de obra en actividades agrícolas"),
    ("gtot_mo_ctrl_pla_prod","gtot_mo_ctrl_pla",   "mano de obra en control de plagas"),
    ("gtot_maq_prod",        "gtot_maq_pp",        "alquiler de maquinaria"),
    ("gtot_fanim_prod",      "gtot_fanim_pp",      "alquiler de fuerza animal"),
    ("gtot_agua_riego_prod", "gtot_agua_riego_pp", "agua para riego"),
    ("gtot_mo_riego_prod",   "gtot_mo_riego_pp",   "mano de obra en labores de riego"),
    ("gtot_met_mip_prod",    "gtot_met_mip_pp",    "manejo integrado de plagas"),
    ("gtot_tpte_prod_prod",  "gtot_tpte_prod_pp",  "transporte de insumos y producción"),
    ("gtot_plagui_q_prod",   "gtot_plagui_q_pp",   "plaguicidas"),
    ("gtot_abo_fert_q_prod", "gtot_abo_fert_q_pp", "abonos y fertilizantes químicos"),
    ("gtot_abo_org_prod",    "gtot_abo_org_pp",    "abono orgánico"),
    ("gtot_otr_prod",        "gtot_otr_pp",        "otros conceptos"),
]
for vname, src, desc in _GTOT_PROD:
    add(grupo="Predio y Parcelas", subgrupo="Costos - Productor", script="10",
        variable_creada=vname, tipo="continua",
        vars_input_encuesta="",
        vars_input_derivadas=src,
        nivel_agregacion="Productor",
        universo_filtros="Todos",
        unidad_medida="soles (S/.)",
        definicion=f"Gasto total del productor en {desc}, agregado sobre todas sus parcelas.",
        time_frame="Campaña agrícola",
        disponibilidad="Línea base y línea seguimiento",
        rol_analitico="Outcome",
        familia_outcome="Económicos",
        cuestionario_ref="",
        notas=f"Agregación por productor: bys Codprod22 post: egen total({src}).")

add(grupo="Predio y Parcelas", subgrupo="Costos Operativos - Productor", script="10",
    variable_creada="gtot_oper_1_prod", tipo="continua",
    vars_input_encuesta="",
    vars_input_derivadas="Componentes de gtot_oper_1_pp agregados al productor",
    nivel_agregacion="Productor",
    universo_filtros="Todos",
    unidad_medida="soles (S/.)",
    definicion="Gasto total del productor en costos operativos, incluyendo insumos químicos y orgánicos, agregado sobre todas sus parcelas.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="",
    notas="")

add(grupo="Predio y Parcelas", subgrupo="Costos Operativos - Productor", script="10",
    variable_creada="gtot_oper_2_prod", tipo="continua",
    vars_input_encuesta="",
    vars_input_derivadas="Componentes de gtot_oper_2_pp agregados al productor",
    nivel_agregacion="Productor",
    universo_filtros="Todos",
    unidad_medida="soles (S/.)",
    definicion="Gasto total del productor en costos operativos, excluyendo insumos químicos y orgánicos, agregado sobre todas sus parcelas.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="",
    notas="")

# ============================================================================
# GRUPO: Económicos - Cultivo Principal (script 09)
# ============================================================================
# Escala productiva
add(grupo="Económicos - Cultivo Principal", subgrupo="Escala Productiva", script="09",
    variable_creada="tot_has_semb_cult_culp", tipo="continua",
    vars_input_encuesta="preg114d, preg114d1, preg114de",
    vars_input_derivadas="",
    nivel_agregacion="Cultivo Principal",
    universo_filtros="Cultivo principal del productor",
    unidad_medida="hectáreas",
    definicion="Total de hectáreas sembradas o instaladas con el cultivo principal del productor durante la campaña agrícola.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="Módulo XI, pregunta 114d",
    notas="Se estandarizan unidades: conversión de m² a hectáreas cuando preg114d1==2; cuando preg114d1==3 se aplica factor preg114de.")

add(grupo="Económicos - Cultivo Principal", subgrupo="Escala Productiva", script="09",
    variable_creada="tot_ud_plnts_culp", tipo="continua",
    vars_input_encuesta="preg114e",
    vars_input_derivadas="",
    nivel_agregacion="Cultivo Principal",
    universo_filtros="Cultivo principal; cultivos permanentes",
    unidad_medida="unidades (plantas)",
    definicion="Total de plantas instaladas del cultivo principal del productor (relevante para cultivos permanentes).",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="Módulo XI, pregunta 114e",
    notas="Relevante para cítricos y plátano; missing para cultivos transitorios como la papa.")

add(grupo="Económicos - Cultivo Principal", subgrupo="Escala Productiva", script="09",
    variable_creada="tot_ud_plnts_eprod_culp", tipo="continua",
    vars_input_encuesta="preg114f",
    vars_input_derivadas="tot_ud_plnts_culp",
    nivel_agregacion="Cultivo Principal",
    universo_filtros="Cultivo principal; cultivos permanentes",
    unidad_medida="unidades (plantas)",
    definicion="Total de plantas en edad productiva del cultivo principal del productor.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="Módulo XI, pregunta 114f",
    notas="Se imputa 0 cuando tot_ud_plnts_culp>0 pero preg114f está missing.")

add(grupo="Económicos - Cultivo Principal", subgrupo="Escala Productiva", script="09",
    variable_creada="tot_kg_sem_culp", tipo="continua",
    vars_input_encuesta="preg114e1, preg114e2, preg114e4",
    vars_input_derivadas="",
    nivel_agregacion="Cultivo Principal",
    universo_filtros="Cultivo principal; cultivos transitorios",
    unidad_medida="kilogramos",
    definicion="Total de semilla utilizada para sembrar el cultivo principal del productor.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="Módulo XI, preguntas 114e1-e4",
    notas="Conversión de toneladas a kilogramos cuando preg114e2==2; uso de preg114e4 cuando preg114e2==3.")

# Producción y ventas
_PROD_VENTAS = [
    ("tot_kg_prod_cose_culp",     "producción cosechada",                                   "Módulo XI, pregunta 114g",  "preg114g, preg114g1"),
    ("tot_kg_prod_vend_culp",     "producción cosechada vendida",                           "Módulo XI, pregunta 114h1", "preg114h1"),
    ("tot_kg_prod_vend_MN_culp",  "producción vendida al mercado nacional",                 "Módulo XI, pregunta 114k",  "preg114k"),
    ("tot_kg_prod_vend_MI_culp",  "producción vendida al mercado internacional",            "Módulo XI, pregunta 114k1", "preg114k1"),
]
for vname, desc, ref, pregs in _PROD_VENTAS:
    add(grupo="Económicos - Cultivo Principal", subgrupo="Producción y Ventas", script="09",
        variable_creada=vname, tipo="continua",
        vars_input_encuesta=pregs,
        vars_input_derivadas="preg114ge (factor de conversión por moda de tipo de cultivo y unidad)",
        nivel_agregacion="Cultivo Principal",
        universo_filtros="Cultivo principal",
        unidad_medida="kilogramos",
        definicion=f"Total de {desc} del cultivo principal del productor durante la campaña agrícola.",
        time_frame="Campaña agrícola",
        disponibilidad="Línea base y línea seguimiento",
        rol_analitico="Outcome",
        familia_outcome="Económicos",
        cuestionario_ref=ref,
        notas="Se estandarizan unidades reportadas: kilogramos, toneladas u otras unidades locales (con factor de conversión preg114ge asignado por moda según tipo de cultivo y unidad).")

# Valor de la producción cosechada (insumo para gasto alquiler %prod en script 10)
add(grupo="Económicos - Cultivo Principal", subgrupo="Valor de la Producción", script="09",
    variable_creada="ixkg_ppc_imp", tipo="continua",
    vars_input_encuesta="preg114h2, preg114h1 (vía ixkg_ppc)",
    vars_input_derivadas="ixkg_ppc; mediana por nomb_tipo_cult × post cuando ixkg_ppc está missing",
    nivel_agregacion="Cultivo",
    universo_filtros="Todos los cultivos",
    unidad_medida="soles (S/.) por kilogramo",
    definicion="Ingreso implícito por kilogramo del cultivo con imputación: cuando el cultivo no fue vendido se asigna la mediana del precio por tipo de cultivo × periodo.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Insumo",
    familia_outcome="Económicos",
    cuestionario_ref="Módulo XI, preguntas 114h1 y 114h2",
    notas="Sólo se usa como precio sombra para valorizar la producción cosechada en vtot_cose_ppc. No reemplaza a ixkg_ppc en otros análisis.")

add(grupo="Económicos - Cultivo Principal", subgrupo="Valor de la Producción", script="09",
    variable_creada="vtot_cose_culp", tipo="continua",
    vars_input_encuesta="",
    vars_input_derivadas="tot_kg_prod_cose_ppc × ixkg_ppc_imp (en frame all_crops, antes del filtro a cultivo principal y rename _ppc→_culp)",
    nivel_agregacion="Cultivo Principal",
    universo_filtros="Cultivo principal",
    unidad_medida="soles (S/.)",
    definicion="Valor total en soles de la producción cosechada del cultivo principal (kilogramos cosechados × precio implícito imputado).",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Insumo",
    familia_outcome="Económicos",
    cuestionario_ref="",
    notas="Insumo para valorizar el pago de alquiler reportado como % de la producción (gtot_alq_b_pp en script 10).")

add(grupo="Predio y Parcelas", subgrupo="Valor de la Producción", script="09",
    variable_creada="vtot_cose_pp", tipo="continua",
    vars_input_encuesta="",
    vars_input_derivadas="vtot_cose_ppc (suma sobre cultivos del predio, en frame all_crops de 09; equivalente a vtot_cose_culp en la base de cultivo principal)",
    nivel_agregacion="Parcela",
    universo_filtros="Todos",
    unidad_medida="soles (S/.)",
    definicion="Valor total en soles de la producción cosechada del predio (suma del valor de los cultivos del predio).",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Insumo",
    familia_outcome="Económicos",
    cuestionario_ref="",
    notas="Exportada en Valor_Produccion_Predio_LByLS.dta. Se consume en 10_build_farm.do para valorizar el pago de alquiler como % de producción (gtot_alq_b_pp).")

# Plaguicidas (cultivo principal)
add(grupo="Económicos - Cultivo Principal", subgrupo="Insumos - Plaguicidas", script="09",
    variable_creada="usa_plag_culp", tipo="binaria",
    vars_input_encuesta="preg114x1_1",
    vars_input_derivadas="",
    nivel_agregacion="Cultivo Principal",
    universo_filtros="Cultivo principal",
    unidad_medida="0/1",
    definicion="El productor usa algún plaguicida (fungicida, herbicida, insecticida o nematicida) en el cultivo principal.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome;Filtro",
    familia_outcome="Económicos",
    cuestionario_ref="Módulo XI, pregunta 114x1_1",
    notas="Cuando usa_plag_culp==0, las variables kg, litros y gasto de plaguicidas se forzan a 0.")

add(grupo="Económicos - Cultivo Principal", subgrupo="Insumos - Plaguicidas", script="09",
    variable_creada="kg_plag_culp", tipo="continua",
    vars_input_encuesta="preg114x2a_1 a preg114x4_5",
    vars_input_derivadas="",
    nivel_agregacion="Cultivo Principal",
    universo_filtros="Cultivo principal",
    unidad_medida="kilogramos",
    definicion="Total de plaguicidas aplicados en el cultivo principal del productor, expresado en kilogramos.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="Módulo XI, preguntas 114x",
    notas="Construida en helper outliers_pesticide_prices.do. Suma sobre hasta 5 productos distintos reportados. Se fija a 0 cuando usa_plag_culp==0.")

add(grupo="Económicos - Cultivo Principal", subgrupo="Insumos - Plaguicidas", script="09",
    variable_creada="lt_plag_culp", tipo="continua",
    vars_input_encuesta="preg114x2a_1 a preg114x4_5",
    vars_input_derivadas="",
    nivel_agregacion="Cultivo Principal",
    universo_filtros="Cultivo principal",
    unidad_medida="litros",
    definicion="Total de plaguicidas aplicados en el cultivo principal del productor, expresado en litros.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="Módulo XI, preguntas 114x",
    notas="Se fija a 0 cuando usa_plag_culp==0.")

add(grupo="Económicos - Cultivo Principal", subgrupo="Insumos - Plaguicidas", script="09",
    variable_creada="kgxha_plag_culp", tipo="continua",
    vars_input_encuesta="",
    vars_input_derivadas="kg_plag_culp, tot_has_semb_cult_culp",
    nivel_agregacion="Cultivo Principal",
    universo_filtros="Cultivo principal",
    unidad_medida="kilogramos por hectárea (kg/ha)",
    definicion="Intensidad de uso de plaguicidas sólidos en el cultivo principal, expresada como kilogramos aplicados por hectárea sembrada.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="Módulo XI, preguntas 114x",
    notas="Calculada solo cuando la superficie sembrada es positiva. Se fija a 0 cuando usa_plag_culp==0.")

add(grupo="Económicos - Cultivo Principal", subgrupo="Insumos - Plaguicidas", script="09",
    variable_creada="ltxha_plag_culp", tipo="continua",
    vars_input_encuesta="",
    vars_input_derivadas="lt_plag_culp, tot_has_semb_cult_culp",
    nivel_agregacion="Cultivo Principal",
    universo_filtros="Cultivo principal",
    unidad_medida="litros por hectárea (lt/ha)",
    definicion="Intensidad de uso de plaguicidas líquidos en el cultivo principal, expresada como litros aplicados por hectárea sembrada.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="Módulo XI, preguntas 114x",
    notas="Calculada solo cuando la superficie sembrada es positiva. Se fija a 0 cuando usa_plag_culp==0.")

add(grupo="Económicos - Cultivo Principal", subgrupo="Insumos - Plaguicidas", script="09",
    variable_creada="gtot_plag_culp", tipo="continua",
    vars_input_encuesta="preg114x",
    vars_input_derivadas="kg_plag_culp, lt_plag_culp",
    nivel_agregacion="Cultivo Principal",
    universo_filtros="Cultivo principal",
    unidad_medida="soles (S/.)",
    definicion="Gasto total en plaguicidas aplicados en el cultivo principal del productor.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="Módulo XI, preguntas 114x",
    notas="Serie en soles corrientes. Ver gtot_plag_culp_def para la versión deflactada.")

add(grupo="Económicos - Cultivo Principal", subgrupo="Insumos - Plaguicidas", script="09",
    variable_creada="gtot_plag_culp_def", tipo="continua",
    vars_input_encuesta="",
    vars_input_derivadas="gtot_plag_culp",
    nivel_agregacion="Cultivo Principal",
    universo_filtros="Cultivo principal",
    unidad_medida="soles constantes de 2021 (S/.)",
    definicion="Gasto total en plaguicidas aplicados en el cultivo principal del productor, expresado en soles constantes de 2021.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="",
    notas="Línea base sin ajuste; línea seguimiento dividida por el factor de deflación (IPC julio-diciembre 2022 / IPC julio-diciembre 2021 ≈ 1.0847). Fuente del IPC: INEI.")

# Abono orgánico
add(grupo="Económicos - Cultivo Principal", subgrupo="Insumos - Abono Orgánico", script="09",
    variable_creada="usa_aboo_culp", tipo="binaria",
    vars_input_encuesta="preg114y1a",
    vars_input_derivadas="",
    nivel_agregacion="Cultivo Principal",
    universo_filtros="Cultivo principal",
    unidad_medida="0/1",
    definicion="El productor usa abono orgánico o natural en el cultivo principal.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome;Filtro",
    familia_outcome="Económicos",
    cuestionario_ref="Módulo XI, pregunta 114y1a",
    notas="")

add(grupo="Económicos - Cultivo Principal", subgrupo="Insumos - Abono Orgánico", script="09",
    variable_creada="kg_aboo_culp", tipo="continua",
    vars_input_encuesta="preg114y1b, preg114y1c, preg114y1e, preg114y1ee",
    vars_input_derivadas="",
    nivel_agregacion="Cultivo Principal",
    universo_filtros="Cultivo principal",
    unidad_medida="kilogramos",
    definicion="Total de abono orgánico o natural aplicado en el cultivo principal del productor, expresado en kilogramos.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="Módulo XI, preguntas 114y",
    notas="Se fija a 0 cuando usa_aboo_culp==0.")

add(grupo="Económicos - Cultivo Principal", subgrupo="Insumos - Abono Orgánico", script="09",
    variable_creada="lt_aboo_culp", tipo="continua",
    vars_input_encuesta="preg114y1b, preg114y1c, preg114y1e, preg114y1ee",
    vars_input_derivadas="",
    nivel_agregacion="Cultivo Principal",
    universo_filtros="Cultivo principal",
    unidad_medida="litros",
    definicion="Total de abono orgánico o natural aplicado en el cultivo principal del productor, expresado en litros.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="Módulo XI, preguntas 114y",
    notas="Valor extremo de 50000 recodificado manualmente a missing. Se fija a 0 cuando usa_aboo_culp==0.")

add(grupo="Económicos - Cultivo Principal", subgrupo="Insumos - Abono Orgánico", script="09",
    variable_creada="gtot_aboo_culp", tipo="continua",
    vars_input_encuesta="preg114y1f",
    vars_input_derivadas="",
    nivel_agregacion="Cultivo Principal",
    universo_filtros="Cultivo principal",
    unidad_medida="soles (S/.)",
    definicion="Gasto total en abono orgánico o natural aplicado en el cultivo principal del productor.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="Módulo XI, pregunta 114y1f",
    notas="Serie en soles corrientes.")

add(grupo="Económicos - Cultivo Principal", subgrupo="Insumos - Abono Orgánico", script="09",
    variable_creada="gtot_aboo_culp_def", tipo="continua",
    vars_input_encuesta="",
    vars_input_derivadas="gtot_aboo_culp",
    nivel_agregacion="Cultivo Principal",
    universo_filtros="Cultivo principal",
    unidad_medida="soles constantes de 2021 (S/.)",
    definicion="Gasto total en abono orgánico aplicado en el cultivo principal del productor, expresado en soles constantes de 2021.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="",
    notas="Línea base sin ajuste; línea seguimiento dividida por el factor de deflación.")

# Abono químico
add(grupo="Económicos - Cultivo Principal", subgrupo="Insumos - Abono Químico", script="09",
    variable_creada="usa_aboq_culp", tipo="binaria",
    vars_input_encuesta="preg114z1a",
    vars_input_derivadas="",
    nivel_agregacion="Cultivo Principal",
    universo_filtros="Cultivo principal",
    unidad_medida="0/1",
    definicion="El productor usa abono químico en el cultivo principal.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome;Filtro",
    familia_outcome="Económicos",
    cuestionario_ref="Módulo XI, pregunta 114z1a",
    notas="")

add(grupo="Económicos - Cultivo Principal", subgrupo="Insumos - Abono Químico", script="09",
    variable_creada="kg_aboq_culp", tipo="continua",
    vars_input_encuesta="preg114z1b, preg114z1c, preg114z1e, preg114z1ee",
    vars_input_derivadas="",
    nivel_agregacion="Cultivo Principal",
    universo_filtros="Cultivo principal",
    unidad_medida="kilogramos",
    definicion="Total de abono químico aplicado en el cultivo principal del productor, expresado en kilogramos.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="Módulo XI, preguntas 114z",
    notas="")

add(grupo="Económicos - Cultivo Principal", subgrupo="Insumos - Abono Químico", script="09",
    variable_creada="lt_aboq_culp", tipo="continua",
    vars_input_encuesta="preg114z1b, preg114z1c, preg114z1e, preg114z1ee",
    vars_input_derivadas="",
    nivel_agregacion="Cultivo Principal",
    universo_filtros="Cultivo principal",
    unidad_medida="litros",
    definicion="Total de abono químico aplicado en el cultivo principal del productor, expresado en litros.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="Módulo XI, preguntas 114z",
    notas="")

add(grupo="Económicos - Cultivo Principal", subgrupo="Insumos - Abono Químico", script="09",
    variable_creada="gtot_aboq_culp", tipo="continua",
    vars_input_encuesta="preg114z1f",
    vars_input_derivadas="",
    nivel_agregacion="Cultivo Principal",
    universo_filtros="Cultivo principal",
    unidad_medida="soles (S/.)",
    definicion="Gasto total en abono químico aplicado en el cultivo principal del productor.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="Módulo XI, pregunta 114z1f",
    notas="Serie en soles corrientes.")

add(grupo="Económicos - Cultivo Principal", subgrupo="Insumos - Abono Químico", script="09",
    variable_creada="gtot_aboq_culp_def", tipo="continua",
    vars_input_encuesta="",
    vars_input_derivadas="gtot_aboq_culp",
    nivel_agregacion="Cultivo Principal",
    universo_filtros="Cultivo principal",
    unidad_medida="soles constantes de 2021 (S/.)",
    definicion="Gasto total en abono químico aplicado en el cultivo principal del productor, expresado en soles constantes de 2021.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="",
    notas="Línea base sin ajuste; línea seguimiento dividida por el factor de deflación.")

# Eventos adversos
add(grupo="Económicos - Cultivo Principal", subgrupo="Eventos Adversos", script="09",
    variable_creada="afct_plga_culp", tipo="binaria",
    vars_input_encuesta="preg114o1",
    vars_input_derivadas="",
    nivel_agregacion="Cultivo Principal",
    universo_filtros="Cultivo principal",
    unidad_medida="0/1",
    definicion="Alguna plaga afectó al cultivo principal del productor durante la campaña agrícola.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="Módulo XI, pregunta 114o1",
    notas="")

add(grupo="Económicos - Cultivo Principal", subgrupo="Eventos Adversos", script="09",
    variable_creada="afct_evi_culp", tipo="binaria",
    vars_input_encuesta="preg114o2",
    vars_input_derivadas="",
    nivel_agregacion="Cultivo Principal",
    universo_filtros="Cultivo principal",
    unidad_medida="0/1",
    definicion="Algún evento inesperado (climático u otro) afectó al cultivo principal del productor durante la campaña agrícola.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="Módulo XI, pregunta 114o2",
    notas="")

# Ingresos, márgenes y rendimientos (versión base)
add(grupo="Económicos - Cultivo Principal", subgrupo="Ingresos y Márgenes", script="09",
    variable_creada="itot_culp", tipo="continua",
    vars_input_encuesta="preg114h2",
    vars_input_derivadas="tot_kg_prod_vend_culp, ixkg_culp",
    nivel_agregacion="Cultivo",
    universo_filtros="Todos los cultivos del productor",
    unidad_medida="soles (S/.)",
    definicion="Ingreso total por la venta del cultivo durante la campaña agrícola.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="Módulo XI, pregunta 114h2",
    notas="Se calcula como el producto de los kilogramos vendidos y el ingreso por kilogramo implícito.")

add(grupo="Económicos - Cultivo Principal", subgrupo="Ingresos y Márgenes", script="09",
    variable_creada="itot_culp_def", tipo="continua",
    vars_input_encuesta="",
    vars_input_derivadas="itot_culp",
    nivel_agregacion="Cultivo",
    universo_filtros="Todos los cultivos del productor",
    unidad_medida="soles constantes de 2021 (S/.)",
    definicion="Ingreso total por la venta del cultivo durante la campaña agrícola, expresado en soles constantes de 2021.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="",
    notas="Versión deflactada usada como outcome en el script 22.")

add(grupo="Económicos - Cultivo Principal", subgrupo="Ingresos y Márgenes", script="09",
    variable_creada="ixkg_culp", tipo="continua",
    vars_input_encuesta="preg114h2",
    vars_input_derivadas="itot_culp, tot_kg_prod_vend_culp",
    nivel_agregacion="Cultivo",
    universo_filtros="Todos los cultivos con ventas positivas",
    unidad_medida="soles por kilogramo (S/./kg)",
    definicion="Ingreso por kilogramo vendido del cultivo durante la campaña agrícola.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Covariable",
    familia_outcome="Económicos",
    cuestionario_ref="",
    notas="")

add(grupo="Económicos - Cultivo Principal", subgrupo="Ingresos y Márgenes", script="09",
    variable_creada="ixha_culp", tipo="continua",
    vars_input_encuesta="",
    vars_input_derivadas="itot_culp, tot_has_semb_cult_culp",
    nivel_agregacion="Cultivo",
    universo_filtros="Todos los cultivos con área sembrada positiva",
    unidad_medida="soles por hectárea (S/./ha)",
    definicion="Ingreso bruto por hectárea sembrada del cultivo durante la campaña agrícola.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Covariable",
    familia_outcome="Económicos",
    cuestionario_ref="",
    notas="")

add(grupo="Económicos - Cultivo Principal", subgrupo="Ingresos y Márgenes", script="09",
    variable_creada="mgn_ins_culp", tipo="continua",
    vars_input_encuesta="",
    vars_input_derivadas="itot_culp, gtot_plag_culp, gtot_aboo_culp, gtot_aboq_culp",
    nivel_agregacion="Cultivo Principal",
    universo_filtros="Cultivo principal",
    unidad_medida="soles (S/.)",
    definicion="Margen bruto sobre el gasto en insumos (plaguicidas, abono orgánico y abono químico) del cultivo principal del productor durante la campaña agrícola.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="",
    notas="Construida como itot_culp menos la suma de los tres gastos en insumos.")

add(grupo="Económicos - Cultivo Principal", subgrupo="Ingresos y Márgenes", script="09",
    variable_creada="mgn_ins_culp_def", tipo="continua",
    vars_input_encuesta="",
    vars_input_derivadas="mgn_ins_culp",
    nivel_agregacion="Cultivo Principal",
    universo_filtros="Cultivo principal",
    unidad_medida="soles constantes de 2021 (S/.)",
    definicion="Margen bruto sobre el gasto en insumos del cultivo principal del productor durante la campaña agrícola, expresado en soles constantes de 2021.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="",
    notas="Versión deflactada usada como outcome en el script 22.")

add(grupo="Económicos - Cultivo Principal", subgrupo="Ingresos y Márgenes", script="09",
    variable_creada="mgn_ins_ha_culp", tipo="continua",
    vars_input_encuesta="",
    vars_input_derivadas="mgn_ins_culp, tot_has_semb_cult_culp",
    nivel_agregacion="Cultivo Principal",
    universo_filtros="Cultivo principal con área sembrada positiva",
    unidad_medida="soles por hectárea (S/./ha)",
    definicion="Margen bruto sobre el gasto en insumos por hectárea sembrada del cultivo principal del productor.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="",
    notas="")

# Rendimientos
add(grupo="Económicos - Cultivo Principal", subgrupo="Rendimientos", script="09",
    variable_creada="kgxha_semb_culp", tipo="continua",
    vars_input_encuesta="",
    vars_input_derivadas="tot_kg_prod_cose_culp, tot_has_semb_cult_culp",
    nivel_agregacion="Cultivo",
    universo_filtros="Cultivos con área sembrada positiva",
    unidad_medida="kilogramos por hectárea (kg/ha)",
    definicion="Rendimiento del cultivo expresado en kilogramos cosechados por hectárea sembrada.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="",
    notas="Cociente entre kilogramos cosechados y hectáreas sembradas.")

add(grupo="Económicos - Cultivo Principal", subgrupo="Rendimientos", script="09",
    variable_creada="kgx1p_eprod_culp", tipo="continua",
    vars_input_encuesta="",
    vars_input_derivadas="tot_kg_prod_cose_culp, tot_ud_plnts_eprod_culp",
    nivel_agregacion="Cultivo",
    universo_filtros="Cultivos con plantas en edad productiva positivas",
    unidad_medida="kilogramos por planta",
    definicion="Rendimiento de la planta en edad productiva, expresado en kilogramos cosechados por planta.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="",
    notas="Relevante para cultivos permanentes.")

add(grupo="Económicos - Cultivo Principal", subgrupo="Rendimientos", script="09",
    variable_creada="kgxkg_sem_culp", tipo="continua",
    vars_input_encuesta="",
    vars_input_derivadas="tot_kg_prod_cose_culp, tot_kg_sem_culp",
    nivel_agregacion="Cultivo",
    universo_filtros="Cultivos con semilla utilizada positiva",
    unidad_medida="kilogramos por kilogramo (kg/kg)",
    definicion="Rendimiento de la semilla, expresado en kilogramos cosechados por kilogramo de semilla utilizado.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Económicos",
    cuestionario_ref="",
    notas="Relevante para cultivos transitorios.")

# Versiones winsorizadas de rendimientos, ingresos y márgenes
_WZ_DESCS = [
    ("wz1", "winsorizada con capeo en logaritmos (capeo 1)"),
    ("wz2", "winsorizada con capeo en niveles (capeo 2)"),
    ("mis", "con valores atípicos reemplazados a missing"),
]
_WZ_BASE = [
    ("kgxha_semb_culp",  "Rendimientos", "Rendimiento del cultivo en kilogramos por hectárea sembrada",                                     "kilogramos por hectárea (kg/ha)"),
    ("kgx1p_eprod_culp", "Rendimientos", "Rendimiento de la planta en edad productiva, en kilogramos cosechados por planta",                "kilogramos por planta"),
    ("kgxkg_sem_culp",   "Rendimientos", "Rendimiento de la semilla, en kilogramos cosechados por kilogramo de semilla utilizado",          "kilogramos por kilogramo (kg/kg)"),
    ("itot_culp",        "Ingresos y Márgenes", "Ingreso total por la venta del cultivo",                                                    "soles (S/.)"),
    ("ixkg_culp",        "Ingresos y Márgenes", "Ingreso por kilogramo vendido del cultivo",                                                 "soles por kilogramo (S/./kg)"),
    ("ixha_culp",        "Ingresos y Márgenes", "Ingreso bruto por hectárea sembrada del cultivo",                                           "soles por hectárea (S/./ha)"),
    ("gtot_plag_culp",  "Insumos - Plaguicidas", "Gasto total en plaguicidas aplicados en el cultivo principal",                       "soles (S/.)"),
    ("mgn_ins_culp",    "Ingresos y Márgenes", "Margen bruto sobre el gasto en insumos del cultivo principal",                         "soles (S/.)"),
    ("mgn_ins_ha_culp", "Ingresos y Márgenes", "Margen bruto sobre el gasto en insumos por hectárea sembrada del cultivo principal",  "soles por hectárea (S/./ha)"),
]
for base, sub, desc_base, unit in _WZ_BASE:
    for suf, suf_desc in _WZ_DESCS:
        vname = f"{base}_{suf}"
        add(grupo="Económicos - Cultivo Principal", subgrupo=sub, script="09",
            variable_creada=vname, tipo="continua",
            vars_input_encuesta="",
            vars_input_derivadas=base,
            nivel_agregacion="Cultivo Principal" if "culp" in base else "Cultivo",
            universo_filtros="Cultivo principal" if "culp" in base else "Todos los cultivos",
            unidad_medida=unit,
            definicion=f"{desc_base}, versión {suf_desc}.",
            time_frame="Campaña agrícola",
            disponibilidad="Línea base y línea seguimiento",
            rol_analitico="Outcome",
            familia_outcome="Económicos",
            cuestionario_ref="",
            notas=f"No se emplea como outcome principal; sirve para análisis de sensibilidad frente a valores atípicos. Se incluye por trazabilidad del catálogo.")

# ============================================================================
# GRUPO: Conocimiento Agronómico (script 11)
# ============================================================================
_TEST_NOTES_STD = "La estandarización within-crop (división por la desviación estándar del grupo control) absorbe las diferencias de escala entre tests."

for cult, full, max_total, max_bpa in [
    ("CQ",  "Cítricos", 17, 12),
    ("PAQ", "Papa",     17, 12),
    ("PLQ", "Plátano",  17, 11),
]:
    add(grupo="Conocimiento Agronómico", subgrupo=f"Test de {full}", script="11",
        variable_creada=f"ptj_{cult}re", tipo="score",
        vars_input_encuesta=f"{cult}1 a {cult}{max_total}",
        vars_input_derivadas="",
        nivel_agregacion="Productor",
        universo_filtros=f"post==1; productores evaluados en {full}",
        unidad_medida=f"puntaje (0-{max_total})",
        definicion=f"Puntaje total del test de conocimiento en buenas prácticas agrícolas aplicado al cultivo de {full.lower()}.",
        time_frame="Línea seguimiento",
        disponibilidad="Solo línea seguimiento",
        rol_analitico="Covariable",
        familia_outcome="",
        cuestionario_ref=f"Test de BPAs en {full}",
        notas="Ponderado por dificultad de pregunta y reescalado al rango original (0-17).")

    add(grupo="Conocimiento Agronómico", subgrupo=f"Test de {full}", script="11",
        variable_creada=f"ptj_{cult}re_std", tipo="score",
        vars_input_encuesta=f"{cult}1 a {cult}{max_total}",
        vars_input_derivadas=f"ptj_{cult}re",
        nivel_agregacion="Productor",
        universo_filtros=f"post==1; productores evaluados en {full}",
        unidad_medida="desviaciones estándar",
        definicion=f"Puntaje total estandarizado del test de conocimiento en buenas prácticas agrícolas aplicado al cultivo de {full.lower()}.",
        time_frame="Línea seguimiento",
        disponibilidad="Solo línea seguimiento",
        rol_analitico="Outcome",
        familia_outcome="Conocimiento",
        cuestionario_ref=f"Test de BPAs en {full}",
        notas=_TEST_NOTES_STD)

    add(grupo="Conocimiento Agronómico", subgrupo=f"Test de {full}", script="11",
        variable_creada=f"ptj_BPA_{cult}", tipo="score",
        vars_input_encuesta=f"{cult}5 a {cult}{max_total}",
        vars_input_derivadas="",
        nivel_agregacion="Productor",
        universo_filtros=f"post==1; productores evaluados en {full}",
        unidad_medida=f"puntaje (0-{max_bpa})",
        definicion=f"Puntaje de la sección de buenas prácticas agrícolas del test de conocimiento aplicado al cultivo de {full.lower()}.",
        time_frame="Línea seguimiento",
        disponibilidad="Solo línea seguimiento",
        rol_analitico="Covariable",
        familia_outcome="",
        cuestionario_ref=f"Test de BPAs en {full}",
        notas=f"Subconjunto de preguntas de BPA del test; escala 0-{max_bpa}.")

    add(grupo="Conocimiento Agronómico", subgrupo=f"Test de {full}", script="11",
        variable_creada=f"ptj_BPA_{cult}_std", tipo="score",
        vars_input_encuesta=f"{cult}5 a {cult}{max_total}",
        vars_input_derivadas=f"ptj_BPA_{cult}",
        nivel_agregacion="Productor",
        universo_filtros=f"post==1; productores evaluados en {full}",
        unidad_medida="desviaciones estándar",
        definicion=f"Puntaje estandarizado de la sección de buenas prácticas agrícolas del test de conocimiento aplicado al cultivo de {full.lower()}.",
        time_frame="Línea seguimiento",
        disponibilidad="Solo línea seguimiento",
        rol_analitico="Outcome",
        familia_outcome="Conocimiento",
        cuestionario_ref=f"Test de BPAs en {full}",
        notas=_TEST_NOTES_STD)

# Pooled
add(grupo="Conocimiento Agronómico", subgrupo="Pooled (Todos los Cultivos)", script="11",
    variable_creada="ptj_test", tipo="score",
    vars_input_encuesta="CQ1-CQ17, PAQ1-PAQ17, PLQ1-PLQ17 (según cultivo)",
    vars_input_derivadas="ptj_CQre, ptj_PAQre, ptj_PLQre",
    nivel_agregacion="Productor",
    universo_filtros="post==1",
    unidad_medida="puntaje (0-17)",
    definicion="Puntaje total del test de conocimiento, consolidado entre los tres cultivos evaluados.",
    time_frame="Línea seguimiento",
    disponibilidad="Solo línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Conocimiento",
    cuestionario_ref="",
    notas="Cada productor responde un único test según el cultivo evaluado. Se obtiene tomando el primer valor no missing entre los tres puntajes específicos.")

add(grupo="Conocimiento Agronómico", subgrupo="Pooled (Todos los Cultivos)", script="11",
    variable_creada="ptj_test_std", tipo="score",
    vars_input_encuesta="CQ1-CQ17, PAQ1-PAQ17, PLQ1-PLQ17",
    vars_input_derivadas="ptj_CQre_std, ptj_PAQre_std, ptj_PLQre_std",
    nivel_agregacion="Productor",
    universo_filtros="post==1",
    unidad_medida="desviaciones estándar",
    definicion="Puntaje total estandarizado del test de conocimiento, consolidado entre los tres cultivos evaluados.",
    time_frame="Línea seguimiento",
    disponibilidad="Solo línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Conocimiento",
    cuestionario_ref="",
    notas="")

add(grupo="Conocimiento Agronómico", subgrupo="Pooled (Todos los Cultivos)", script="11",
    variable_creada="ptj_BPA", tipo="score",
    vars_input_encuesta="CQ5-CQ17, PAQ5-PAQ17, PLQ5-PLQ17",
    vars_input_derivadas="ptj_BPA_CQ, ptj_BPA_PAQ, ptj_BPA_PLQ",
    nivel_agregacion="Productor",
    universo_filtros="post==1",
    unidad_medida="puntaje (0-11 o 0-12 según cultivo)",
    definicion="Puntaje de la sección de buenas prácticas agrícolas del test de conocimiento, consolidado entre los tres cultivos evaluados.",
    time_frame="Línea seguimiento",
    disponibilidad="Solo línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Conocimiento",
    cuestionario_ref="",
    notas="La escala varía entre cultivos: 0-12 para cítricos y papa, 0-11 para plátano.")

add(grupo="Conocimiento Agronómico", subgrupo="Pooled (Todos los Cultivos)", script="11",
    variable_creada="ptj_BPA_std", tipo="score",
    vars_input_encuesta="CQ5-CQ17, PAQ5-PAQ17, PLQ5-PLQ17",
    vars_input_derivadas="ptj_BPA_CQ_std, ptj_BPA_PAQ_std, ptj_BPA_PLQ_std",
    nivel_agregacion="Productor",
    universo_filtros="post==1",
    unidad_medida="desviaciones estándar",
    definicion="Puntaje estandarizado de la sección de buenas prácticas agrícolas del test de conocimiento, consolidado entre los tres cultivos evaluados.",
    time_frame="Línea seguimiento",
    disponibilidad="Solo línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Conocimiento",
    cuestionario_ref="",
    notas="")

# ============================================================================
# GRUPO: BPA - Prácticas Agronómicas (script 12)
# ============================================================================
_BPA_MAIN = [
    ("bpa_1",  "Suelo",                           "preg4011", "El productor realiza análisis de suelos."),
    ("bpa_2",  "Suelo",                           "preg4012", "El productor mezcla tierra con materia orgánica."),
    ("bpa_3",  "Suelo",                           "preg4013", "El productor asocia cultivos para proteger el suelo."),
    ("bpa_4",  "Suelo",                           "preg4014", "El productor realiza surcos en contornos."),
    ("bpa_5",  "Riego",                           "preg4015", "El productor determina la necesidad de agua del cultivo."),
    ("bpa_6",  "Riego",                           "preg4016", "El productor determina la frecuencia de riego."),
    ("bpa_7",  "Riego",                           "preg4017", "El productor mide la cantidad de agua aplicada al cultivo."),
    ("bpa_8",  "Riego",                           "preg4018", "El productor realiza mantenimiento del sistema de riego."),
    ("bpa_9",  "Riego",                           "preg4019", "El productor analiza el agua de riego."),
    ("bpa_10", "Insumos - Abonos",                "preg4020", "El productor usa abonos en sus cultivos."),
    ("bpa_11", "Insumos - Fertilizantes",         "preg4021", "El productor usa fertilizantes en sus cultivos."),
    ("bpa_12", "Insumos - Plaguicidas",           "preg4022", "El productor usa plaguicidas en sus cultivos."),
    ("bpa_13", "Insumos - Control Biológico",     "preg4023", "El productor aplica control biológico en sus cultivos."),
    ("bpa_14", "Insumos - Manejo Integrado",      "preg4024", "El productor aplica manejo integrado de plagas en sus cultivos."),
]
for v, sub, preg, defn in _BPA_MAIN:
    fam = "BPA no condicionadas" if v in [f"bpa_{i}" for i in range(1,10)] else "BPA condicionadas"
    add(grupo="BPA - Prácticas Agronómicas", subgrupo=sub, script="12",
        variable_creada=v, tipo="binaria",
        vars_input_encuesta=preg,
        vars_input_derivadas="",
        nivel_agregacion="Productor",
        universo_filtros="Todos",
        unidad_medida="0/1",
        definicion=defn,
        time_frame="Campaña agrícola",
        disponibilidad="Línea base y línea seguimiento",
        rol_analitico="Outcome",
        familia_outcome=fam,
        cuestionario_ref=f"Módulo IV, {preg}",
        notas="Las variables bpa_10 a bpa_14 actúan además como condicionantes para las sub-prácticas (bpa_10_*, bpa_11_*, etc.).")

# Sub-prácticas condicionadas
_BPA_SUB = [
    ("bpa_10_1", "Insumos - Abonos",       "preg402",  "El productor usa la cantidad necesaria de abono.",                         "bpa_10==1"),
    ("bpa_10_2", "Insumos - Abonos",       "preg403",  "El productor usa abono de buena calidad.",                                  "bpa_10==1"),
    ("bpa_10_3", "Insumos - Abonos",       "preg404",  "El productor usa abono recomendado por un especialista.",                   "bpa_10==1"),
    ("bpa_10_4", "Insumos - Abonos",       "preg405",  "El productor aplica la dosis recomendada de abono.",                        "bpa_10==1 & bpa_10_3==1"),
    ("bpa_10_5", "Insumos - Abonos",       "preg406",  "El productor almacena el abono en las condiciones recomendadas.",           "bpa_10==1 & bpa_10_3==1"),
    ("bpa_11_1", "Insumos - Fertilizantes","preg407",  "El productor usa la cantidad necesaria de fertilizante.",                   "bpa_11==1"),
    ("bpa_11_2", "Insumos - Fertilizantes","preg408",  "El productor usa fertilizante de buena calidad.",                           "bpa_11==1"),
    ("bpa_11_3", "Insumos - Fertilizantes","preg409",  "El productor usa fertilizante recomendado por un especialista.",            "bpa_11==1"),
    ("bpa_11_4", "Insumos - Fertilizantes","preg410",  "El productor aplica la dosis recomendada de fertilizante.",                 "bpa_11==1 & bpa_11_3==1"),
    ("bpa_11_5", "Insumos - Fertilizantes","preg411",  "El productor almacena el fertilizante en las condiciones recomendadas.",    "bpa_11==1 & bpa_11_3==1"),
    ("bpa_12_1", "Insumos - Plaguicidas",  "preg412",  "El productor usa la cantidad necesaria de plaguicida.",                     "bpa_12==1"),
    ("bpa_12_2", "Insumos - Plaguicidas",  "preg413",  "El productor usa plaguicida de buena calidad.",                              "bpa_12==1"),
    ("bpa_12_3", "Insumos - Plaguicidas",  "preg414",  "El productor usa plaguicida recomendado por un especialista.",              "bpa_12==1"),
    ("bpa_12_4", "Insumos - Plaguicidas",  "preg415",  "El productor usa plaguicida químico.",                                       "bpa_12==1"),
    ("bpa_12_q_1",  "Insumos - Plaguicidas Químicos", "preg415a",          "El productor usa plaguicida químico con etiqueta en el envase.",                                                                         "bpa_12_4==1"),
    ("bpa_12_q_2",  "Insumos - Plaguicidas Químicos", "preg416",           "El productor lee la información impresa en el envase del plaguicida químico.",                                                           "bpa_12_q_1==1"),
    ("bpa_12_q_3",  "Insumos - Plaguicidas Químicos", "preg417",           "El productor aplica la dosis de plaguicida químico recomendada en el envase.",                                                           "bpa_12_q_1==1"),
    ("bpa_12_q_4",  "Insumos - Plaguicidas Químicos", "preg418",           "El productor aplica el plaguicida químico solo en el cultivo indicado en el envase.",                                                    "bpa_12_q_1==1"),
    ("bpa_12_q_5",  "Insumos - Plaguicidas Químicos", "preg419",           "El productor cumple el tiempo mínimo recomendado entre la última aplicación del plaguicida químico y la cosecha (periodo de carencia).", "bpa_12_4==1"),
    ("bpa_12_q_6",  "Insumos - Plaguicidas Químicos", "preg420",           "El productor almacena el plaguicida químico en las condiciones recomendadas.",                                                           "bpa_12_4==1"),
    ("bpa_12_q_7",  "Insumos - Plaguicidas Químicos", "preg4211-preg4216", "El productor usa al menos un equipo de protección personal al aplicar el plaguicida químico.",                                          "bpa_12_4==1"),
    ("bpa_12_q_71", "Insumos - Plaguicidas",          "preg4221-preg4226", "El productor realiza buena gestión de los envases vacíos del plaguicida (usa solo métodos adecuados y no recurre a métodos inadecuados).", "!mi(preg4221)"),
    ("bpa_13_1", "Insumos - Control Biológico","preg423", "El productor realiza evaluación de plagas al aplicar control biológico.","bpa_13==1"),
    ("bpa_14_1", "Insumos - Manejo Integrado", "preg4241-preg4245", "El productor combina de forma estratégica distintos tipos de control dentro del manejo integrado de plagas.", "bpa_14==1"),
]
for v, sub, preg, defn, filt in _BPA_SUB:
    add(grupo="BPA - Prácticas Agronómicas", subgrupo=sub, script="12",
        variable_creada=v, tipo="binaria",
        vars_input_encuesta=preg,
        vars_input_derivadas=filt.split(" & ")[0].split("==")[0].strip() if "==" in filt else "",
        nivel_agregacion="Productor",
        universo_filtros=filt,
        unidad_medida="0/1",
        definicion=defn,
        time_frame="Campaña agrícola",
        disponibilidad="Línea base y línea seguimiento",
        rol_analitico="Outcome",
        familia_outcome="BPA condicionadas",
        cuestionario_ref=f"Módulo IV, {preg}",
        notas="Práctica condicionada a las variables indicadas en la columna universo_filtros.")

# ============================================================================
# GRUPO: Registros y Almacenamiento (script 13)
# ============================================================================
_REGIS = [
    ("regis_1", "preg400R1", "El productor tiene registro o libreta de aplicación de abonos o fertilizantes."),
    ("regis_2", "preg400R2", "El productor tiene registro o libreta de aplicación de plaguicidas."),
    ("regis_3", "preg400R3", "El productor tiene registro o libreta de liberación o aplicación de control biológico."),
    ("regis_4", "preg400R4", "El productor tiene registro o libreta de aplicación de riego."),
    ("regis_5", "preg400R5", "El productor tiene registro de control del producto cosechado."),
    ("regis_6", "preg400R6", "El productor tiene un kardex para controlar el ingreso y salida del almacén."),
    ("regis_7", "preg400R7", "El productor tiene registro de rastreabilidad y trazabilidad."),
]
for v, preg, defn in _REGIS:
    add(grupo="Registros y Almacenamiento", subgrupo="Registros Agrícolas", script="13",
        variable_creada=v, tipo="binaria",
        vars_input_encuesta=preg,
        vars_input_derivadas="",
        nivel_agregacion="Productor",
        universo_filtros="Todos",
        unidad_medida="0/1",
        definicion=defn,
        time_frame="Campaña agrícola",
        disponibilidad="Línea base y línea seguimiento",
        rol_analitico="Outcome",
        familia_outcome="Registros e Inocuidad",
        cuestionario_ref=f"Módulo IV, {preg}",
        notas="")

add(grupo="Registros y Almacenamiento", subgrupo="Condiciones de Almacenamiento", script="13",
    variable_creada="tot_cond_min_alm", tipo="conteo",
    vars_input_encuesta="preg400R8_11, preg400R8_12, preg400R8_13, preg400R8_15, preg400R8_111, preg400R8_112",
    vars_input_derivadas="",
    nivel_agregacion="Productor",
    universo_filtros="Todos",
    unidad_medida="cantidad (0-6)",
    definicion="Total de condiciones mínimas indispensables para el almacenamiento de agroquímicos que cumple el productor.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Registros e Inocuidad",
    cuestionario_ref="Módulo IV, preguntas 400R8",
    notas="Suma de las seis condiciones críticas del almacén.")

# ============================================================================
# GRUPO: Inocuidad Alimentaria (script 14)
# ============================================================================
_INO = [
    ("ino_resid_cult_v1", "Residuos de Cultivos",    "preg5011-preg5015",  "flexible",        "El productor aplica al menos una buena práctica de manejo de residuos de cultivos."),
    ("ino_resid_cult_v2", "Residuos de Cultivos",    "preg5011-preg5015",  "original ENA",    "El productor aplica al menos una buena práctica de manejo de residuos de cultivos y no incurre en prácticas de quema ni botado."),
    ("ino_resid_cult_v3", "Residuos de Cultivos",    "preg5011-preg5015",  "estricta",        "El productor aplica al menos una buena práctica de manejo de residuos de cultivos y no incurre en prácticas de quema, botado ni dejar residuos en campo."),
    ("ino_resid_anim_v1", "Residuos de Animales",    "preg502b1-preg502b5","flexible",        "El productor aplica al menos una buena práctica de manejo de residuos animales."),
    ("ino_resid_anim_v2", "Residuos de Animales",    "preg502b1-preg502b5","original ENA",    "El productor aplica al menos una buena práctica de manejo de residuos animales y no incurre en prácticas de quema ni botado."),
    ("ino_resid_anim_v3", "Residuos de Animales",    "preg502b1-preg502b5","estricta",        "El productor aplica al menos una buena práctica de manejo de residuos animales y no incurre en prácticas de quema, botado ni dejar residuos en campo."),
    ("ino_alim_prod_v1",  "Almacenamiento Alimentos","preg5061-preg5066",  "flexible",        "El productor aplica al menos una buena práctica de almacenamiento de alimentos producidos."),
    ("ino_alim_prod_v2",  "Almacenamiento Alimentos","preg5061-preg5066",  "original ENA",    "El productor aplica al menos una buena práctica de almacenamiento de alimentos y evita al menos una mala práctica."),
    ("ino_alim_prod_v3",  "Almacenamiento Alimentos","preg5061-preg5066",  "estricta",        "El productor aplica al menos una buena práctica de almacenamiento de alimentos y evita todas las malas prácticas consideradas."),
]
for v, sub, preg, ver, defn in _INO:
    add(grupo="Inocuidad Alimentaria", subgrupo=sub, script="14",
        variable_creada=v, tipo="binaria",
        vars_input_encuesta=preg,
        vars_input_derivadas="",
        nivel_agregacion="Productor",
        universo_filtros=f"!mi({preg.split('-')[0]})",
        unidad_medida="0/1",
        definicion=defn,
        time_frame="Campaña agrícola",
        disponibilidad="Línea base y línea seguimiento",
        rol_analitico="Outcome",
        familia_outcome="Registros e Inocuidad",
        cuestionario_ref=f"Módulo V, {preg}",
        notas=f"Versión {ver}. Consumida por el compuesto ENA con la misma rigurosidad.")

add(grupo="Inocuidad Alimentaria", subgrupo="Sensibilización", script="14",
    variable_creada="ino_info_conta_alim", tipo="binaria",
    vars_input_encuesta="preg504",
    vars_input_derivadas="",
    nivel_agregacion="Productor",
    universo_filtros="!mi(preg504)",
    unidad_medida="0/1",
    definicion="Alguna entidad del Estado informó al productor sobre la contaminación de alimentos.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Registros e Inocuidad",
    cuestionario_ref="Módulo V, pregunta 504",
    notas="")

add(grupo="Inocuidad Alimentaria", subgrupo="Etiquetado", script="14",
    variable_creada="ino_etiq_alim", tipo="binaria",
    vars_input_encuesta="preg507",
    vars_input_derivadas="",
    nivel_agregacion="Productor",
    universo_filtros="!mi(preg507)",
    unidad_medida="0/1",
    definicion="El productor identifica o etiqueta los alimentos que produce para su consumo.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Registros e Inocuidad",
    cuestionario_ref="Módulo V, pregunta 507",
    notas="")

add(grupo="Inocuidad Alimentaria", subgrupo="Certificación", script="14",
    variable_creada="ino_cert_cal", tipo="binaria",
    vars_input_encuesta="preg508",
    vars_input_derivadas="",
    nivel_agregacion="Productor",
    universo_filtros="!mi(preg508)",
    unidad_medida="0/1",
    definicion="Los alimentos producidos por el productor cuentan con certificación de calidad.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Registros e Inocuidad",
    cuestionario_ref="Módulo V, pregunta 508",
    notas="")

# ============================================================================
# GRUPO: Compuesto Propio (script 15)
# ============================================================================
_COMP_PROPIO_PILAR = [
    ("bpa_comp_suelo",   "bpa_1, bpa_2, bpa_3, bpa_4",
        "El productor implementa al menos una de las cuatro buenas prácticas agrícolas de suelo."),
    ("bpa_comp_riego",   "bpa_5, bpa_6, bpa_7, bpa_8, bpa_9",
        "El productor implementa al menos una de las cinco buenas prácticas agrícolas de riego."),
    ("bpa_comp_insumos", "bpa_12, bpa_13, bpa_14",
        "El productor no usa plaguicidas, o aplica control biológico, o aplica manejo integrado de plagas."),
    ("bpa_comp_prod",    "bpa_comp_suelo, bpa_comp_riego, bpa_comp_insumos",
        "El productor cumple al menos uno de los tres subpilares del pilar de producción (suelo, riego o insumos)."),
]
for v, deriv, defn in _COMP_PROPIO_PILAR:
    add(grupo="Compuesto Propio BPA", subgrupo="Pilar Producción", script="15",
        variable_creada=v, tipo="binaria",
        vars_input_encuesta="preg4011-preg4024",
        vars_input_derivadas=deriv,
        nivel_agregacion="Productor",
        universo_filtros="Todos",
        unidad_medida="0/1",
        definicion=defn,
        time_frame="Campaña agrícola",
        disponibilidad="Línea base y línea seguimiento",
        rol_analitico="Covariable",
        familia_outcome="",
        cuestionario_ref="Módulo IV",
        notas="Pilar intermedio que alimenta el compuesto final implementa_bpa.")

for ver, suf, src in [("vF","flexible","ino_resid_cult_v1"), ("vO","original","ino_resid_cult_v2"), ("vE","estricta","ino_resid_cult_v3")]:
    add(grupo="Compuesto Propio BPA", subgrupo="Pilar Higiene", script="15",
        variable_creada=f"bpa_comp_higiene_{ver}", tipo="binaria",
        vars_input_encuesta="preg5011-preg5015",
        vars_input_derivadas=src,
        nivel_agregacion="Productor",
        universo_filtros=f"!mi({src})",
        unidad_medida="0/1",
        definicion=f"El productor cumple el pilar de higiene en su versión {suf}.",
        time_frame="Campaña agrícola",
        disponibilidad="Línea base y línea seguimiento",
        rol_analitico="Covariable",
        familia_outcome="",
        cuestionario_ref="Módulo V",
        notas=("Versión consumida por el compuesto final implementa_bpa." if ver=="vO" else f"Versión {suf}, disponible como alternativa para análisis de sensibilidad."))

for ver, suf, src in [("vF","flexible","ino_alim_prod_v1"), ("vO","original","ino_alim_prod_v2"), ("vE","estricta","ino_alim_prod_v3")]:
    add(grupo="Compuesto Propio BPA", subgrupo="Pilar Almacén", script="15",
        variable_creada=f"bpa_comp_alm_{ver}", tipo="binaria",
        vars_input_encuesta="preg5061-preg5066, preg400R6",
        vars_input_derivadas=f"{src}, regis_6",
        nivel_agregacion="Productor",
        universo_filtros="Todos",
        unidad_medida="0/1",
        definicion=f"El productor cumple el pilar de almacenamiento en su versión {suf}.",
        time_frame="Campaña agrícola",
        disponibilidad="Línea base y línea seguimiento",
        rol_analitico="Covariable",
        familia_outcome="",
        cuestionario_ref="Módulo V; Módulo IV, pregunta 400R6",
        notas=("Versión consumida por el compuesto final implementa_bpa." if ver=="vO" else f"Versión {suf}, disponible como alternativa para análisis de sensibilidad."))

add(grupo="Compuesto Propio BPA", subgrupo="Pilar Distribución", script="15",
    variable_creada="bpa_comp_distrib", tipo="binaria",
    vars_input_encuesta="preg400R7, preg507",
    vars_input_derivadas="regis_7, ino_etiq_alim",
    nivel_agregacion="Productor",
    universo_filtros="Todos",
    unidad_medida="0/1",
    definicion="El productor implementa trazabilidad o etiquetado de los alimentos que produce.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Covariable",
    familia_outcome="",
    cuestionario_ref="Módulo IV, pregunta 400R7; Módulo V, pregunta 507",
    notas="")

add(grupo="Compuesto Propio BPA", subgrupo="Compuesto Final", script="15",
    variable_creada="implementa_bpa", tipo="binaria",
    vars_input_encuesta="preg4011-preg4024, preg5011-preg5015, preg5061-preg5066, preg400R6, preg400R7, preg507",
    vars_input_derivadas="bpa_comp_prod, bpa_comp_higiene_vO, bpa_comp_alm_vO, bpa_comp_distrib",
    nivel_agregacion="Productor",
    universo_filtros="Todos",
    unidad_medida="0/1",
    definicion="El productor implementa de manera conjunta los cuatro pilares de buenas prácticas agrícolas: producción, higiene, almacenamiento y distribución.",
    time_frame="Campaña agrícola",
    disponibilidad="Línea base y línea seguimiento",
    rol_analitico="Outcome",
    familia_outcome="Compuesto Propio",
    cuestionario_ref="Módulos IV y V",
    notas="Intersección (AND) estricta de los cuatro pilares. Se emplean las versiones intermedias vO de higiene y almacenamiento.")

# ============================================================================
# GRUPO: Compuesto ENA (script 15)
# ============================================================================
_ENA_SUBS = [
    ("riego",      "preg4015-preg4019", "bpa_5, bpa_6, bpa_7, bpa_8, bpa_9",
        {"vF": "El productor cumple al menos dos de las cinco buenas prácticas de riego.",
         "vO": "El productor cumple simultáneamente las prácticas de medición de agua aplicada y mantenimiento del sistema, y al menos una entre determinación de necesidad hídrica, frecuencia de riego o análisis del agua.",
         "vE": "El productor cumple las cinco buenas prácticas de riego consideradas."}),
    ("suelo",      "preg4011-preg4014", "bpa_1, bpa_2, bpa_3, bpa_4",
        {"vF": "El productor cumple al menos dos de las cuatro buenas prácticas de suelo.",
         "vO": "El productor realiza análisis de suelos y al menos una entre: mezcla con materia orgánica, asociación de cultivos o surcos en contornos.",
         "vE": "El productor cumple las cuatro buenas prácticas de suelo consideradas."}),
    ("fert_abo",   "preg402-preg411, preg400R1", "bpa_10, bpa_11, bpa_10_3, bpa_11_3, regis_1",
        {"vF": "El productor usa abono o fertilizante y adicionalmente lleva registro de aplicación o usa el producto recomendado.",
         "vO": "El productor usa abono o fertilizante recomendado por un especialista y lleva registro de aplicación.",
         "vE": "El productor cumple todas las sub-prácticas del uso de abono y fertilizante y lleva registro de aplicación."}),
    ("plag",       "preg412-preg420, preg4211-preg4216, preg4221-preg4226, preg400R2", "bpa_12, bpa_12_4, bpa_12_q_1, bpa_12_q_2, bpa_12_q_3, bpa_12_q_4, bpa_12_q_5, bpa_12_q_6, bpa_12_q_7, regis_2",
        {"vF": "El productor usa plaguicidas y adicionalmente lleva registro o usa el producto recomendado. Si el plaguicida es químico, cumple al menos una práctica de manejo seguro.",
         "vO": "El productor usa plaguicidas, lleva registro y usa el producto recomendado. Si el plaguicida es químico, además revisa la etiqueta y cumple al menos dos de cinco prácticas de manejo seguro.",
         "vE": "El productor cumple todas las sub-prácticas del uso de plaguicidas y lleva registro. Si el plaguicida es químico, cumple además las cinco prácticas de manejo seguro consideradas."}),
    ("biocontrol", "preg423, preg400R3", "bpa_13, bpa_13_1, regis_3",
        {"vF": "El productor aplica control biológico y adicionalmente evalúa plagas o lleva registro.",
         "vO": "El productor aplica control biológico, evalúa plagas y lleva registro.",
         "vE": "El productor aplica control biológico, evalúa plagas y lleva registro."}),
    ("mip",        "preg4241-preg4245", "bpa_14, bpa_14_1",
        {"vF": "El productor aplica manejo integrado de plagas y combina distintos métodos de control.",
         "vO": "El productor aplica manejo integrado de plagas y combina distintos métodos de control.",
         "vE": "El productor aplica manejo integrado de plagas y combina distintos métodos de control."}),
    ("inoc",       "preg5011-preg5015, preg504, preg5061-preg5066, preg507, preg400R7", "ino_info_conta_alim, ino_etiq_alim, regis_7",
        {"vF": "El productor implementa al menos una práctica de prevención (manejo de residuos de cultivo en versión flexible o haber recibido información sobre contaminación de alimentos) y al menos una práctica de control (almacenamiento de alimentos en versión flexible, etiquetado o trazabilidad).",
         "vO": "El productor implementa al menos una práctica de prevención (manejo de residuos de cultivo en versión original o información sobre contaminación) y al menos una práctica de control (almacenamiento en versión original, etiquetado o trazabilidad).",
         "vE": "El productor implementa al menos una práctica de prevención (manejo de residuos de cultivo en versión estricta o información sobre contaminación) y al menos una práctica de control (almacenamiento en versión estricta, etiquetado o trazabilidad)."}),
]
_VER_LBL = {"vF":"flexible","vO":"original ENA","vE":"estricta"}

# Mapeo de version de inocuidad a derivadas específicas
_INOC_DERIV = {"vF": "ino_resid_cult_v1, ino_info_conta_alim, ino_alim_prod_v1, ino_etiq_alim, regis_7",
               "vO": "ino_resid_cult_v2, ino_info_conta_alim, ino_alim_prod_v2, ino_etiq_alim, regis_7",
               "vE": "ino_resid_cult_v3, ino_info_conta_alim, ino_alim_prod_v3, ino_etiq_alim, regis_7"}

for sub, pregs, deriv_base, defns in _ENA_SUBS:
    for ver in ("vF", "vO", "vE"):
        vname = f"bpa_ena_{sub}_{ver}"
        deriv = _INOC_DERIV[ver] if sub == "inoc" else deriv_base
        add(grupo="Compuesto ENA", subgrupo=f"Subindicador - {sub}", script="15",
            variable_creada=vname, tipo="binaria",
            vars_input_encuesta=pregs,
            vars_input_derivadas=deriv,
            nivel_agregacion="Productor",
            universo_filtros="Todos",
            unidad_medida="0/1",
            definicion=defns[ver],
            time_frame="Campaña agrícola",
            disponibilidad="Línea base y línea seguimiento",
            rol_analitico="Outcome",
            familia_outcome="Compuesto ENA",
            cuestionario_ref="Módulo IV" if sub != "inoc" else "Módulo V",
            notas=f"Versión {_VER_LBL[ver]}. El subindicador de inocuidad consume la variante de la cadena correspondiente a su versión (vF→v1, vO→v2, vE→v3).")

# Pilares ENA
_ENA_PILARES = [
    ("agro",    "bpa_ena_riego_{ver}, bpa_ena_suelo_{ver}",
        "El productor cumple el pilar agronómico en su versión {suf} (sub-indicador de riego o de suelo)."),
    ("insumos", "bpa_ena_fert_abo_{ver}, bpa_ena_plag_{ver}, bpa_ena_biocontrol_{ver}, bpa_ena_mip_{ver}",
        "El productor cumple el pilar de insumos en su versión {suf} (al menos un sub-indicador entre fertilización/abono, plaguicidas, control biológico o manejo integrado de plagas)."),
    ("inoc",    "bpa_ena_inoc_{ver}",
        "El productor cumple el pilar de inocuidad en su versión {suf}."),
]
for sub, deriv_tmpl, defn_tmpl in _ENA_PILARES:
    for ver in ("vF", "vO", "vE"):
        vname = f"ena_pilar_{sub}_{ver}"
        add(grupo="Compuesto ENA", subgrupo=f"Pilar - {sub}", script="15",
            variable_creada=vname, tipo="binaria",
            vars_input_encuesta="preg4011-preg424" if sub != "inoc" else "preg5011-preg507, preg400R7",
            vars_input_derivadas=deriv_tmpl.format(ver=ver),
            nivel_agregacion="Productor",
            universo_filtros="Todos",
            unidad_medida="0/1",
            definicion=defn_tmpl.format(suf=_VER_LBL[ver], ver=ver),
            time_frame="Campaña agrícola",
            disponibilidad="Línea base y línea seguimiento",
            rol_analitico="Outcome",
            familia_outcome="Compuesto ENA",
            cuestionario_ref="Módulo IV" if sub != "inoc" else "Módulo V",
            notas=f"Versión {_VER_LBL[ver]}.")

for ver in ("vF", "vO", "vE"):
    add(grupo="Compuesto ENA", subgrupo="Compuesto Final", script="15",
        variable_creada=f"implementa_bpa_ena_{ver}", tipo="binaria",
        vars_input_encuesta="preg4011-preg424, preg5011-preg507, preg400R1-preg400R7",
        vars_input_derivadas=f"ena_pilar_agro_{ver}, ena_pilar_insumos_{ver}, ena_pilar_inoc_{ver}",
        nivel_agregacion="Productor",
        universo_filtros="Todos",
        unidad_medida="0/1",
        definicion=f"El productor cumple al menos dos de los tres pilares del compuesto ENA (agronómico, insumos e inocuidad) en su versión {_VER_LBL[ver]}.",
        time_frame="Campaña agrícola",
        disponibilidad="Línea base y línea seguimiento",
        rol_analitico="Outcome",
        familia_outcome="Compuesto ENA",
        cuestionario_ref="Módulos IV y V",
        notas=f"Versión {_VER_LBL[ver]}. El criterio de agregación es común a las tres versiones; las diferencias provienen de las definiciones de sub-indicadores.")

# ============================================================================
# HOJA LEYENDA
# ============================================================================
LEYENDA_ROWS = [
    ("Campo", "Valor", "Definición"),
    ("tipo", "binaria", "Variable dicotómica 0/1."),
    ("tipo", "continua", "Variable numérica continua."),
    ("tipo", "categórica", "Variable con categorías discretas no ordinales (códigos)."),
    ("tipo", "conteo", "Variable de conteo discreta no negativa."),
    ("tipo", "índice", "Variable compuesta construida por agregación ponderada."),
    ("tipo", "score", "Puntaje de un test (eventualmente estandarizado)."),
    ("nivel_agregacion", "Productor", "Una observación por productor (Codprod22)."),
    ("nivel_agregacion", "Productor o Jefe de Hogar", "Información imputada del roster del hogar cuando no hay coincidencia directa con el productor."),
    ("nivel_agregacion", "Hogar", "Una observación por hogar."),
    ("nivel_agregacion", "Parcela", "Una observación por productor-parcela."),
    ("nivel_agregacion", "Cultivo", "Una observación por productor-predio-cultivo (todos los cultivos)."),
    ("nivel_agregacion", "Cultivo Principal", "Una observación por productor-predio-cultivo, filtrando al cultivo principal."),
    ("nivel_agregacion", "Centro Poblado", "Una observación por centro poblado."),
    ("disponibilidad", "Solo línea base", "Variable disponible únicamente en la encuesta de línea base."),
    ("disponibilidad", "Solo línea seguimiento", "Variable disponible únicamente en la encuesta de línea seguimiento."),
    ("disponibilidad", "Línea base y línea seguimiento", "Variable disponible en ambas rondas."),
    ("rol_analitico", "Outcome", "Variable usada como dependiente en al menos una regresión del pipeline."),
    ("rol_analitico", "Covariable", "Variable usada como control en regresiones o como descriptor."),
    ("rol_analitico", "Filtro", "Variable usada en condiciones if para definir universos o submuestras."),
    ("familia_outcome", "Conocimiento", "Puntajes de test de conocimiento agronómico (script 18)."),
    ("familia_outcome", "BPA no condicionadas", "BPAs generales no condicionadas al cultivo (script 19)."),
    ("familia_outcome", "BPA condicionadas", "BPAs condicionadas al uso de insumos (script 20)."),
    ("familia_outcome", "Registros e Inocuidad", "Registros agrícolas y prácticas de inocuidad alimentaria (script 21)."),
    ("familia_outcome", "Económicos", "Escala productiva, producción, ventas, insumos, márgenes (script 22)."),
    ("familia_outcome", "Compuesto Propio", "Compuesto de BPA propio (intersección estricta de cuatro pilares)."),
    ("familia_outcome", "Compuesto ENA", "Subindicadores, pilares y compuesto ENA (scripts 25A y 25B)."),
    ("familia_outcome", "Test Falsificación", "Variable usada en el test de falsificación del script 16."),
]

# ============================================================================
# HOJA PENDIENTES
# ============================================================================
PENDIENTES_ROWS = [
    ("variable", "razón pendiente"),
    ("kg_plag_culp, lt_plag_culp",
     "Construidas dentro del helper outliers_pesticide_prices.do en el frame plaguicds. Queda pendiente revisar con más detalle la lógica exacta de detección de outliers."),
    ("Versiones _wz1, _wz2, _mis (script 09)",
     "Incluidas en el catálogo para trazabilidad; no se usan como outcomes principales pero quedan disponibles para análisis de sensibilidad."),
    ("Variables sea1-sea23 y cap_suelo, cap_riego, cap_insum (script 08)",
     "Rol analítico por definir. Pendiente decidir su uso en análisis descriptivos o de heterogeneidad."),
    ("Ingresos del hogar (imen_hog, imen_per_cap, log_imen_per_cap, imen_dep_hog, imen_indep_hog, imen_otr_hog)",
     "Rol analítico por definir. Pendiente decidir su uso en análisis descriptivos."),
    ("Gastos agregados a nivel de parcela o productor (gtot_*_pp, gtot_*_prod)",
     "Reclasificadas como outcomes de la familia Económicos. Confirmar especificaciones de las estimaciones ITT/LATE asociadas."),
]

# ============================================================================
# CONSTRUCCION DEL EXCEL
# ============================================================================
def bold_fill(cell, fill_hex=BID_BLUE, font_color="FFFFFF"):
    cell.font = Font(bold=True, color=font_color, size=11, name="Calibri")
    cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def thin_border(cell):
    side = Side(style="thin", color="BFBFBF")
    cell.border = Border(left=side, right=side, top=side, bottom=side)

wb = openpyxl.Workbook()

# ---------- Hoja 1: Diccionario ----------
ws = wb.active
ws.title = "Diccionario"

# Escribir headers (azul institucional del BID)
for j, h in enumerate(HEADERS, start=1):
    c = ws.cell(row=1, column=j, value=h)
    bold_fill(c)
    thin_border(c)

# Escribir filas (color pastel por grupo temático + notas en lenguaje plano)
for i, row in enumerate(ROWS, start=2):
    grupo = row.get("grupo", "")
    var_name = row.get("variable_creada", "")
    color = GROUP_COLORS.get(grupo, "FFFFFF")
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for j, h in enumerate(HEADERS, start=1):
        val = row[h]
        if h == "notas":
            val = clean_notas(var_name, val)
        c = ws.cell(row=i, column=j, value=val)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.font = Font(size=10, name="Calibri")
        thin_border(c)
        c.fill = fill

# Anchos de columnas
col_widths = {
    "grupo": 24, "subgrupo": 24, "script": 8, "variable_creada": 26, "tipo": 11,
    "vars_input_encuesta": 30, "vars_input_derivadas": 32, "nivel_agregacion": 18,
    "universo_filtros": 26, "unidad_medida": 22, "definicion": 60, "time_frame": 22,
    "disponibilidad": 22, "rol_analitico": 16, "familia_outcome": 22,
    "cuestionario_ref": 26, "notas": 50
}
for j, h in enumerate(HEADERS, start=1):
    ws.column_dimensions[get_column_letter(j)].width = col_widths.get(h, 18)

ws.row_dimensions[1].height = 30
ws.freeze_panes = "E2"
ws.auto_filter.ref = ws.dimensions

# ---------- Hoja 2: Leyenda ----------
ws2 = wb.create_sheet("Leyenda")
for i, row in enumerate(LEYENDA_ROWS, start=1):
    for j, val in enumerate(row, start=1):
        c = ws2.cell(row=i, column=j, value=val)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        thin_border(c)
        if i == 1:
            bold_fill(c)
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 32
ws2.column_dimensions["C"].width = 80
ws2.freeze_panes = "A2"

# ---------- Hoja 3: Pendientes ----------
ws3 = wb.create_sheet("Pendientes")
for i, row in enumerate(PENDIENTES_ROWS, start=1):
    for j, val in enumerate(row, start=1):
        c = ws3.cell(row=i, column=j, value=val)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        thin_border(c)
        if i == 1:
            bold_fill(c)
ws3.column_dimensions["A"].width = 40
ws3.column_dimensions["B"].width = 95
ws3.freeze_panes = "A2"

OUT = r"E:\Consultorías\BID\HRC0052956\1_Data\Diccionarios\Diccionario_Variables_Creadas.xlsx"
wb.save(OUT)

from collections import Counter
cnt_g = Counter(r["grupo"] for r in ROWS)
cnt_s = Counter(r["script"] for r in ROWS)
print(f"[OK] Archivo guardado: {OUT}")
print(f"Total variables documentadas: {len(ROWS)}")
print("\nDesglose por grupo (temática):")
for g, n in sorted(cnt_g.items()):
    print(f"  {g}: {n}")
print("\nDesglose por script:")
for s, n in sorted(cnt_s.items()):
    print(f"  Script {s}: {n}")
print(f"\nPendientes listados: {len(PENDIENTES_ROWS)-1}")
