# -*- coding: utf-8 -*-
"""
Genera Catalogo_Variables_EI_ECAs.xlsx, una version del catalogo pensada para
hacedores de politica (funcionarios BID, SENASA, MIDAGRI): lenguaje plano, sin
jerga Stata, acronimos explicitos y con filas coloreadas por grupo tematico.

Entrada : 1_Data/Diccionarios/Diccionario_Variables_Creadas.xlsx
Salida  : 1_Data/Diccionarios/Catalogo_Variables_EI_ECAs.xlsx

Cambios respecto al diccionario interno:
- Solo 11 campos: Grupo, Subgrupo, Variable creada, Tipo, Nivel de agregacion,
  Unidad de medida, Definicion, Marco temporal, Disponibilidad, Rol analitico, Notas.
- Encabezados con primera letra en mayuscula y sin guiones bajos.
- Colores de relleno por grupo tematico (fila completa) — ver GROUP_COLORS en
  _notas_publicas.
- 'Notas' reescrito para eliminar sintaxis Stata y nombres internos de variables,
  y para explicitar acronimos (ENA, BPA, ECA, MIP, BPM, ITT, LATE, IPC).
- 26 variables excluidas (intensidades por hectarea, rendimientos e ingresos
  unitarios). Ver EXCLUDE_BASE abajo.

Autor : Carlos Marena (asistido)
Fecha : 2026-04-20
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from _notas_publicas import BID_BLUE, GROUP_COLORS, clean_notas

IN_PATH  = r"E:\Consultorías\BID\HRC0052956\1_Data\Diccionarios\Diccionario_Variables_Creadas.xlsx"
OUT_PATH = r"E:\Consultorías\BID\HRC0052956\1_Data\Diccionarios\Catálogo_Variables_EI_ECAs.xlsx"

# Campos a conservar y su etiqueta publica
KEEP_FIELDS = [
    ("grupo",             "Grupo"),
    ("subgrupo",          "Subgrupo"),
    ("variable_creada",   "Variable creada"),
    ("tipo",              "Tipo"),
    ("nivel_agregacion",  "Nivel de agregación"),
    ("unidad_medida",     "Unidad de medida"),
    ("definicion",        "Definición"),
    ("time_frame",        "Marco temporal"),
    ("disponibilidad",    "Disponibilidad"),
    ("rol_analitico",     "Rol analítico"),
    ("notas",             "Notas"),
]

#===============================================================================
# Variables a EXCLUIR del catalogo para policy makers:
# - Intensidades por hectarea (kg/ha, lt/ha, S/./ha)
# - Rendimientos (kg/ha, kg/planta, kg/kg semilla)
# - Ingresos unitarios (S/./kg)
# Estas variables siguen existiendo en el diccionario interno para trazabilidad;
# se omiten aqui porque sus ratios pueden confundir a un publico no tecnico y
# se pueden reconstruir facilmente a partir de las variables en niveles.
#===============================================================================
EXCLUDE_BASE = {
    "kgxha_plag_culp", "ltxha_plag_culp",
    "ixkg_culp", "ixha_culp",
    "mgn_ins_ha_culp",
    "kgxha_semb_culp", "kgx1p_eprod_culp", "kgxkg_sem_culp",
}
# Tambien sus versiones winsorizadas (_wz1, _wz2, _mis)
EXCLUDE_VARS = set(EXCLUDE_BASE)
for base in EXCLUDE_BASE:
    for suf in ("_wz1", "_wz2", "_mis"):
        EXCLUDE_VARS.add(f"{base}{suf}")


#===============================================================================
# Limpieza de 'definicion' especifica del catalogo (expande acronimos)
#===============================================================================
def clean_definicion(var_name: str, definicion: str) -> str:
    """Expande acronimos comunes en la definicion (primera mencion)."""
    if not definicion:
        return ""
    d = definicion
    replacements = [
        ("la ECA",  "la ECA (Escuela de Campo Agrícola)"),
        ("una ECA", "una ECA (Escuela de Campo Agrícola)"),
        ("ECAs",    "ECAs (Escuelas de Campo Agrícolas)"),
        ("BPAs",    "BPAs (Buenas Prácticas Agrícolas)"),
        ("ENA ",    "ENA (Encuesta Nacional Agropecuaria del INEI) "),
        ("MIP",     "MIP (Manejo Integrado de Plagas)"),
        ("BPM",     "BPM (Buenas Prácticas de Manufactura)"),
    ]
    for old, new in replacements:
        if old in d and new.split("(")[0].strip() not in d:
            d = d.replace(old, new, 1)
    return d


#===============================================================================
# Cargar diccionario interno
#===============================================================================
wb_in = openpyxl.load_workbook(IN_PATH, data_only=True)
ws_in = wb_in["Diccionario"]
in_headers = [c.value for c in ws_in[1]]
idx = {h: i for i, h in enumerate(in_headers)}

raw_rows = []
for row in ws_in.iter_rows(min_row=2, values_only=True):
    raw_rows.append({h: (row[idx[h]] if row[idx[h]] is not None else "") for h in in_headers})


#===============================================================================
# Transformar filas (aplicar exclusion + limpiezas)
#===============================================================================
out_rows = []
n_excluded = 0
for r in raw_rows:
    v = r["variable_creada"]
    if v in EXCLUDE_VARS:
        n_excluded += 1
        continue
    out = {}
    for k_in, k_out in KEEP_FIELDS:
        out[k_out] = r.get(k_in, "")
    out["Definición"] = clean_definicion(v, r.get("definicion", ""))
    out["Notas"]      = clean_notas(v, r.get("notas", ""))
    out_rows.append(out)


#===============================================================================
# Escribir libro de salida
#===============================================================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Catálogo"

thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Encabezados
headers_out = [h for _, h in KEEP_FIELDS]
for j, h in enumerate(headers_out, start=1):
    c = ws.cell(row=1, column=j, value=h)
    c.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    c.fill = PatternFill(start_color=BID_BLUE, end_color=BID_BLUE, fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

# Filas de datos
for i, row in enumerate(out_rows, start=2):
    color = GROUP_COLORS.get(row["Grupo"], "FFFFFF")
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for j, h in enumerate(headers_out, start=1):
        c = ws.cell(row=i, column=j, value=row[h])
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = border
        c.fill = fill
        c.font = Font(size=10, name="Calibri")

# Anchos de columna
col_widths = {
    "Grupo":                 26,
    "Subgrupo":              26,
    "Variable creada":       26,
    "Tipo":                  12,
    "Nivel de agregación":   20,
    "Unidad de medida":      22,
    "Definición":            62,
    "Marco temporal":        24,
    "Disponibilidad":        24,
    "Rol analítico":         16,
    "Notas":                 62,
}
for j, h in enumerate(headers_out, start=1):
    ws.column_dimensions[get_column_letter(j)].width = col_widths.get(h, 20)

ws.row_dimensions[1].height = 36
ws.freeze_panes = "D2"
ws.auto_filter.ref = ws.dimensions


#===============================================================================
# Hoja 2: Leyenda (acronimos y terminos)
#===============================================================================
ws2 = wb.create_sheet("Leyenda")

leyenda = [
    ["Término / Acrónimo", "Significado"],
    ["ECA / ECAs", "Escuela(s) de Campo Agrícola(s) — programa de capacitación participativa implementado por SENASA dentro del proyecto PRODESA."],
    ["BPA / BPAs", "Buena(s) Práctica(s) Agrícola(s) — conjunto de prácticas de manejo de suelo, riego, insumos, registros y poscosecha recomendadas para una producción segura y sostenible."],
    ["BPM", "Buenas Prácticas de Manufactura — prácticas que garantizan condiciones sanitarias e higiénicas en el manejo poscosecha y almacenamiento."],
    ["MIP", "Manejo Integrado de Plagas — estrategia de control de plagas que combina medidas biológicas, culturales y químicas con criterio de mínimo impacto."],
    ["ENA", "Encuesta Nacional Agropecuaria del INEI. Fuente de catálogos oficiales de indicadores agropecuarios utilizados como referencia."],
    ["INEI", "Instituto Nacional de Estadística e Informática del Perú."],
    ["SENASA", "Servicio Nacional de Sanidad Agraria del Perú, ejecutor del programa de ECAs."],
    ["IPC", "Índice de Precios al Consumidor — usado para deflactar montos monetarios entre línea base y línea de seguimiento."],
    ["Línea base / Línea de seguimiento", "Encuestas aplicadas a los productores antes de la intervención (línea base) y después (línea de seguimiento)."],
    ["Campaña agrícola", "Periodo productivo anual al que corresponden las variables de producción, insumos y gastos (campaña 2020-2021 en línea base, 2022-2023 en línea de seguimiento)."],
    ["ITT", "Intent-to-Treat (Intención de tratar) — efecto estimado comparando a quienes fueron asignados al programa con quienes no lo fueron, independientemente de si participaron efectivamente."],
    ["LATE", "Local Average Treatment Effect — efecto promedio del tratamiento sobre los productores cuya participación efectivamente cambió por estar asignados al programa."],
    ["Centro poblado / CCPP", "Unidad geográfica de aleatorización del programa. Los centros poblados fueron sorteados a tratamiento o control."],
    ["Productor", "Productor agrícola encuestado (unidad principal de análisis)."],
    ["Cultivo principal", "Cultivo al que el productor destina mayor superficie durante la campaña agrícola; sobre él se recogen las variables más detalladas."],
    ["Parcela", "Unidad física de producción dentro del predio del productor (un productor puede tener varias parcelas)."],
    ["Tipo: continua", "Variable numérica con un rango continuo (ej. hectáreas, kilogramos, soles)."],
    ["Tipo: binaria", "Variable que toma solo dos valores: 1 (sí) y 0 (no)."],
    ["Tipo: categórica", "Variable que toma un conjunto finito de categorías (ej. niveles educativos)."],
    ["Tipo: conteo", "Variable entera no negativa que cuenta ocurrencias (ej. número de miembros del hogar)."],
    ["Tipo: score / índice", "Variable numérica construida como puntaje agregado a partir de varias preguntas."],
    ["Rol: Outcome", "Variable de resultado cuyo impacto busca ser estimado."],
    ["Rol: Covariable", "Variable de control utilizada en los modelos de estimación."],
    ["Rol: Filtro", "Variable que restringe el universo de análisis de otra variable (condiciona su cálculo)."],
    ["Versiones winsorizadas (sufijos _wz1, _wz2, _mis)", "Tratamientos de valores extremos: _wz1 reemplaza por el percentil 99; _wz2 por el 95; _mis los convierte en valor faltante."],
    ["Versiones deflactadas (sufijo _def)", "Montos monetarios expresados en soles constantes de 2021, aplicando un factor de deflación basado en el IPC del INEI."],
]

for i, row in enumerate(leyenda, start=1):
    for j, val in enumerate(row, start=1):
        c = ws2.cell(row=i, column=j, value=val)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = border
        if i == 1:
            c.font = Font(bold=True, color="FFFFFF", size=11)
            c.fill = PatternFill(start_color=BID_BLUE, end_color=BID_BLUE, fill_type="solid")
        else:
            c.font = Font(size=10)

ws2.column_dimensions["A"].width = 36
ws2.column_dimensions["B"].width = 100
ws2.row_dimensions[1].height = 28
ws2.freeze_panes = "A2"


#===============================================================================
# Hoja 3: Indice de grupos (con conteo y color)
#===============================================================================
ws3 = wb.create_sheet("Índice de grupos")
from collections import Counter
cnt = Counter(r["Grupo"] for r in out_rows)

group_desc = {
    "Características de la Observación": "Identificadores de estrato, unidad de aleatorización y momento de la encuesta. Controles estructurales en las regresiones.",
    "Sociodemográficas del Productor": "Edad, sexo, educación, lengua materna y autoidentificación étnica del productor o jefe de hogar.",
    "Hogar - Activos y Condiciones de Vida": "Índices sintéticos de condiciones de vida y tenencia de activos durables.",
    "Hogar - Composición Demográfica": "Número de miembros del hogar y su composición por grupos de edad.",
    "Hogar - Servicios de Extensión Agraria": "Participación en servicios de extensión (asistencia técnica, capacitaciones, visitas) y prácticas agrícolas autorreportadas.",
    "Predio y Parcelas": "Tamaño del predio, número de parcelas, gastos operativos agregados al productor (alquiler, mano de obra, agua, transporte, etc.).",
    "Económicos - Cultivo Principal": "Resultados económicos del cultivo principal: superficie, producción, ventas, ingresos, insumos (plaguicidas, fertilizantes, abono orgánico), eventos adversos, márgenes.",
    "Conocimiento Agronómico": "Puntajes del test de conocimiento aplicado al productor (conocimientos generales, prácticas agronómicas, plagas y enfermedades).",
    "BPA - Prácticas Agronómicas": "Indicadores binarios de adopción de Buenas Prácticas Agrícolas en suelo, riego, fertilización, abono orgánico y plaguicidas.",
    "Registros y Almacenamiento": "Uso de registros (aplicación, producción, gestión) y prácticas de almacenamiento de insumos.",
    "Inocuidad Alimentaria": "Prácticas de higiene, almacenamiento y distribución que minimizan riesgos de contaminación del producto.",
    "Compuesto Propio BPA": "Indicador compuesto construido para esta evaluación que integra cuatro pilares (producción, higiene, almacenamiento, distribución) en un único indicador de cumplimiento.",
    "Compuesto ENA": "Indicador compuesto alineado con el catálogo oficial de la ENA (Encuesta Nacional Agropecuaria del INEI). Tres versiones: Flexible, Original ENA y Estricta.",
}

ws3.cell(row=1, column=1, value="Grupo").font = Font(bold=True, color="FFFFFF")
ws3.cell(row=1, column=1).fill = PatternFill(start_color=BID_BLUE, end_color=BID_BLUE, fill_type="solid")
ws3.cell(row=1, column=2, value="Descripción").font = Font(bold=True, color="FFFFFF")
ws3.cell(row=1, column=2).fill = PatternFill(start_color=BID_BLUE, end_color=BID_BLUE, fill_type="solid")
ws3.cell(row=1, column=3, value="N° variables").font = Font(bold=True, color="FFFFFF")
ws3.cell(row=1, column=3).fill = PatternFill(start_color=BID_BLUE, end_color=BID_BLUE, fill_type="solid")
for c in ws3[1]:
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

seen = []
for r in out_rows:
    if r["Grupo"] not in seen:
        seen.append(r["Grupo"])

for i, g in enumerate(seen, start=2):
    color = GROUP_COLORS.get(g, "FFFFFF")
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for j, val in enumerate([g, group_desc.get(g, ""), cnt[g]], start=1):
        c = ws3.cell(row=i, column=j, value=val)
        c.alignment = Alignment(vertical="top", wrap_text=True,
                                horizontal="center" if j == 3 else "left")
        c.border = border
        c.fill = fill
        c.font = Font(size=10)

ws3.column_dimensions["A"].width = 36
ws3.column_dimensions["B"].width = 90
ws3.column_dimensions["C"].width = 14
ws3.row_dimensions[1].height = 28
ws3.freeze_panes = "A2"


#===============================================================================
# Guardar
#===============================================================================
wb.save(OUT_PATH)
print(f"[OK] Archivo guardado: {OUT_PATH}")
print(f"Total variables en el catálogo: {len(out_rows)}")
print(f"Variables excluidas (intensidades/rendimientos): {n_excluded}")
print(f"\nDesglose por grupo:")
for g in seen:
    print(f"  {g}: {cnt[g]}")
